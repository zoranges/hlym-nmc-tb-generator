"""TB for LCP (Section C) - VMI Generator — Standalone Streamlit App.

严格按 SOP 文档 "TB for LCP" (354行) 实现的 LCP/VMI 零件 TB 自动生成系统.

SOP 情况:
  SITUATION 1: 无颜色 VMI 零件组 → 从 QCPS-Sub Assy 找工序 → 加 -S1 代表行
  SITUATION 2: 有颜色 VMI 零件组 → 从 CIT Matrix 展开色号 → 加 -S1

核心流程 (与 YMC/YMAC 完全不同):
  1. 从 QCPS-Sub Assy 识别 VMI 工序 (STATION = EXTERNAL VENDOR 或供应商名)
  2. 按 PROCESS_NAME 分组, 收集所有 PART_NO
  3. 到 PCL 中查找每个 PART_NO 的 SEC/LVL/PART_NAME/QTY/CU_FA
  4. 判断是否有颜色 (CIT Matrix)
  5. 构建 VMI TB 行:
     - 主件行: PART_NO-S1, PART_NAME SUB ASSY., SUPPLIER=供应商, USE IN=F3
     - 子件行: 原 PART_NO/PART_NAME, SUPPLIER=原供应商, USE IN=供应商名
  6. 如有颜色: 每种颜色展开一组完整的 VMI 行

启动:
    streamlit run lcp_app.py
"""
from __future__ import annotations

import io
import re
import tempfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

import pandas as pd
import streamlit as st
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parent.parent
STRUCT_PATH = PROJECT_ROOT / "NMC Project" / "[POC] TB Structure Explanation.xlsx"

# 内置 VMI 供应商名 (非通用 STATION 值即视为供应商)
_VENDOR_BLACKLIST = {
    "SUB ASSY", "ASSY", "F2", "F2 SUB ASSY", "F2 SUB ASSY & F3",
    "F3", "QTY", "STATION", "EXTERNAL VENDOR / SUB ASSY",
    "HICOM",  # HICOM 是另一个 VMI 供应商, 保留
}
# HICOM 也属于 VMI 供应商 (KOIKE, HICOM, etc.)

# ---------------------------------------------------------------------------
# 行数据结构
# ---------------------------------------------------------------------------

@dataclass
class TBRow:
    sec: str = ""
    lvl1: str = ""
    lvl2: str = ""
    lvl3: str = ""
    lvl4: str = ""
    part_no: str = ""
    part_name: str = ""
    qty: str = ""
    supplier: str = ""
    use_in: str = ""
    ex_new: str = ""
    remarks: str = ""


@dataclass
class VmiGroup:
    """一个 VMI 工序 × 一种颜色 = 一组 TB 行."""
    process_name: str = ""
    vendor_name: str = ""
    color_mark: str = ""   # 空=无颜色
    rows: list[TBRow] = field(default_factory=list)


# ---------------------------------------------------------------------------
# 工具函数
# ---------------------------------------------------------------------------

def _s(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and pd.isna(v):
        return ""
    s = str(v).strip()
    if s.lower() in ("none", "nan"):
        return ""
    return s


# ---------------------------------------------------------------------------
# Part 1: 数据加载
# ---------------------------------------------------------------------------

def _read_sheet(path: Path, sheet_name: str, header_row: int = 0) -> pd.DataFrame:
    df = pd.read_excel(path, sheet_name=sheet_name, header=header_row, engine="openpyxl")
    df.columns = [str(c).strip() if not pd.isna(c) else c for c in df.columns]
    df = df.dropna(how="all").reset_index(drop=True)
    return df


def _find_sheet(wb, keyword: str) -> str | None:
    kw = keyword.lower()
    for sn in wb.sheetnames:
        if kw in sn.lower():
            return sn
    return None


def load_pcl(struct_path: Path) -> pd.DataFrame:
    df = _read_sheet(struct_path, "1. PCL")
    rename = {
        "PARTS SET NO1": "PARTS_SET_NO", "PARTS_SET_NO1": "PARTS_SET_NO",
        "PART NO.": "PART_NO", "PART NAME": "PART_NAME", "QTY": "QUANTITY",
    }
    df = df.rename(columns={k: v for k, v in rename.items() if k in df.columns})
    if "PART_NO" in df.columns:
        df["PART_NO"] = df["PART_NO"].astype(str).str.strip()
        df = df[df["PART_NO"].notna() & (df["PART_NO"] != "") & (df["PART_NO"] != "nan")].reset_index(drop=True)
    if "CU_FA" in df.columns:
        df["CU_FA"] = df["CU_FA"].astype(str).str.strip()
    if "QUANTITY" in df.columns:
        df["QUANTITY"] = df["QUANTITY"].astype(str).str.strip()
    return df


def load_cu_fa_code(struct_path: Path) -> dict[str, str]:
    try:
        df = _read_sheet(struct_path, "6. CU_FA CODE")
        rename = {"Customer Code": "CU_FA_CODE", "G-SU name (FROM)": "SUPPLIER_FROM"}
        df = df.rename(columns={k: v for k, v in rename.items() if k in df.columns})
        if "CU_FA_CODE" not in df.columns or "SUPPLIER_FROM" not in df.columns:
            return {}
        short_name = {
            "YMC": "YMC", "VIETNAM(YMVN)": "YMVN", "IYM(SJP)": "IYM",
            "YIMM(JKT)": "YIMM", "YIMM(WJ)": "YIMM", "TYM B/D-ASSY": "TYM",
            "YMCN": "YMCN", "TWN(YMTT) C BODY": "YMTT", "H.Y.M.M.": "HYMM",
            "H.L.Y.M.": "HLYM", "YPMV": "YPMV", "YMAC": "YMAC",
        }
        result = {}
        for _, row in df.dropna(subset=["CU_FA_CODE"]).drop_duplicates(subset=["CU_FA_CODE"]).iterrows():
            code = str(row["CU_FA_CODE"]).strip()
            full = str(row["SUPPLIER_FROM"]).strip()
            result[code] = short_name.get(full, full)
        return result
    except Exception:
        return {}


def load_ln_bom(struct_path: Path) -> set[str]:
    try:
        df = _read_sheet(struct_path, "7. LN BOM")
        candidates = [c for c in df.columns if str(c).strip().lower() in ("item", "part_no", "part no.")]
        col = candidates[0] if candidates else df.columns[2]
        parts = set()
        for v in df[col].astype(str).str.strip():
            if v and len(v) > 3 and v.lower() != "nan":
                parts.add(v)
        return parts
    except Exception:
        return set()


def load_qcps_sub(struct_path: Path) -> pd.DataFrame:
    """加载 QCPS Sub Assy, 返回含 STATION / PROCESS_NAME / PART_NO 的 DataFrame."""
    wb = load_workbook(struct_path, read_only=True)
    sheet_name = _find_sheet(wb, "sub assy")
    wb.close()
    if not sheet_name:
        return pd.DataFrame(columns=["STATION", "PROCESS_NAME", "PART_NO"])

    wb = load_workbook(struct_path, data_only=True)
    ws = wb[sheet_name]
    max_r, max_c = ws.max_row, ws.max_column
    grid = [[ws.cell(r, c).value for c in range(1, max_c + 1)] for r in range(1, max_r + 1)]
    for rng in ws.merged_cells.ranges:
        anchor = grid[rng.min_row - 1][rng.min_col - 1]
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                grid[r - 1][c - 1] = anchor
    wb.close()

    records = []
    current_station = ""
    current_process = ""
    in_parts = False
    part_col = None

    for row in grid:
        cells_upper = [str(v).upper().strip() if v is not None else "" for v in row]
        if "STATION" in cells_upper and "PROCESS NAME" in cells_upper:
            in_parts = False
            st_idx = cells_upper.index("STATION")
            pn_idx = cells_upper.index("PROCESS NAME")
            ri = grid.index(row)
            if ri + 1 < len(grid):
                nxt = grid[ri + 1]
                current_station = _s(nxt[st_idx]).upper()
                current_process = _s(nxt[pn_idx])
            continue
        if "PART NO" in cells_upper:
            part_col = cells_upper.index("PART NO")
            in_parts = True
            continue
        if in_parts and part_col is not None:
            pn = _s(row[part_col])
            if pn:
                records.append({
                    "STATION": current_station,
                    "PROCESS_NAME": current_process,
                    "PART_NO": pn,
                })

    return pd.DataFrame(records, columns=["STATION", "PROCESS_NAME", "PART_NO"])


def load_cit_matrix(struct_path: Path) -> dict[str, list[dict]]:
    """加载 11. CIT Matrix, 返回 {Part_no: [{color_instruct_no, color_mark}, ...]}.

    CIT Matrix 结构:
      Row 1-4: 多层表头 (MT-code, model, color type, color data)
      Row 5:   列标题 (SEC, Selection, Part no., Part name, Part remarks,
               color instruct no.1, color no.1, color mark1, ...)
      Row 6+:  数据行

    每个 MT-code 对应 5 列: color instruct no., color no., color mark,
                              mak. ind. dwg. prt. no., color remarks
    """
    try:
        wb = load_workbook(struct_path, data_only=True)
        ws = wb["11. CIT Matrix"]

        # 读取第 5 行作为列标题
        headers = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(5, c).value
            headers.append(str(v).strip() if v else f"Col{c}")

        # 找出颜色组: 从 col 5 开始, 每 5 列一组
        # 找到第一个 "color instruct no.X" 列的位置
        color_start = None
        for i, h in enumerate(headers):
            if "color instruct" in h.lower():
                color_start = i
                break

        if color_start is None:
            wb.close()
            return {}

        # 收集所有颜色组
        color_groups = []  # [(instruct_col_idx, color_col_idx, mark_col_idx), ...]
        i = color_start
        while i + 2 < len(headers):
            if "color instruct" in headers[i].lower():
                color_groups.append({
                    "instruct_idx": i,
                    "color_idx": i + 1,
                    "mark_idx": i + 2,
                })
                i += 5
            else:
                break

        # 找到 Part no. 列
        part_col = None
        for i, h in enumerate(headers):
            if h.lower() in ("part no.", "part no"):
                part_col = i
                break

        if part_col is None:
            wb.close()
            return {}

        # 读取数据行 (从第 6 行开始)
        result: dict[str, list[dict]] = {}
        for r in range(6, ws.max_row + 1):
            pn = _s(ws.cell(r, part_col + 1).value)
            if not pn:
                continue

            variants = []
            prev_instruct = ""
            for cg in color_groups:
                ci = _s(ws.cell(r, cg["instruct_idx"] + 1).value)
                cm = _s(ws.cell(r, cg["mark_idx"] + 1).value)

                # 处理继承: "<-----" 表示沿用前一个 color instruct no.
                if ci == "<-----":
                    ci = prev_instruct
                # 跳过 GRAPHIC 标记和空值
                if ci in ("*", "-", "") or ci == "<-----":
                    continue

                prev_instruct = ci

                if ci and cm:
                    # 去重
                    if not any(e["color_instruct_no"] == ci for e in variants):
                        variants.append({"color_instruct_no": ci, "color_mark": cm})

            if variants:
                result[pn] = variants

        wb.close()
        return result
    except Exception:
        return {}


def load_email(_struct_path: Path) -> dict[str, str]:
    return {}


# ---------------------------------------------------------------------------
# Part 2: VMI 工序识别 & 分组
# ---------------------------------------------------------------------------

def _is_vendor_station(station: str) -> bool:
    """判断 STATION 值是否为 VMI 供应商."""
    s = station.upper().strip()
    if not s:
        return False
    if s in {"SUB ASSY", "ASSY", "F2", "F3", "QTY", "STATION", ""}:
        return False
    if s.startswith("F2") or s.startswith("EXTERNAL VENDOR /"):
        return False
    if s.isdigit():
        return False
    # EXTERNAL VENDOR 是通用 VMI 标记, 需要额外获取供应商名
    if s == "EXTERNAL VENDOR":
        return True
    # 含 EXTERNAL VENDOR 的组合
    if "EXTERNAL VENDOR" in s:
        return True
    # 其他值 (KOIKE, HICOM 等) 直接就是供应商名
    return True


def _extract_vendor_name(station: str, process_name: str) -> str:
    """从 STATION 或 PROCESS_NAME 提取供应商名."""
    s = station.upper().strip()
    # 直接是供应商名
    if s not in ("EXTERNAL VENDOR", "EXTERNAL VENDOR & F3", "SUB ASSY", "ASSY", ""):
        if not s.startswith("F2") and not s.isdigit():
            if "EXTERNAL VENDOR" not in s:
                return station.strip()
    # EXTERNAL VENDOR 类: 尝试从 PROCESS_NAME 推断
    # 或使用已知映射 (MODERNRIA 是常见 VMI 供应商)
    return f"VMI:{process_name.strip()[:30]}"


def identify_vmi_processes(qcps_sub: pd.DataFrame) -> pd.DataFrame:
    """在 QCPS-Sub Assy 中识别 VMI 工序.

    返回 DataFrame: PROCESS_NAME, STATION, VENDOR_NAME, 每个 PART_NO 一行
    """
    if qcps_sub.empty:
        return pd.DataFrame(columns=["PROCESS_NAME", "STATION", "VENDOR_NAME", "PART_NO"])

    result = []
    for _, row in qcps_sub.iterrows():
        station = str(row.get("STATION", "")).strip()
        if _is_vendor_station(station):
            process = str(row.get("PROCESS_NAME", "")).strip()
            vendor = _extract_vendor_name(station, process)
            result.append({
                "PROCESS_NAME": process,
                "STATION": station,
                "VENDOR_NAME": vendor,
                "PART_NO": str(row.get("PART_NO", "")).strip(),
            })

    return pd.DataFrame(result, columns=["PROCESS_NAME", "STATION", "VENDOR_NAME", "PART_NO"])


# ---------------------------------------------------------------------------
# Part 3: VMI 组构建
# ---------------------------------------------------------------------------

def _format_subassy_name(name: str) -> str:
    if not name:
        return ""
    name = name.strip()
    upper = name.upper()
    if "SUB ASSY" in upper:
        return name
    if "ASSY" in upper:
        idx = upper.index("ASSY")
        return name[:idx] + "SUB " + name[idx:]
    return f"{name} SUB ASSY."


def _lookup_pcl(part_no: str, pcl: pd.DataFrame) -> dict:
    """在 PCL 中查找零件, 返回 {SEC, LVL1-4, PART_NAME, QTY, CU_FA}."""
    if pcl.empty or "PART_NO" not in pcl.columns:
        return {}
    mask = pcl["PART_NO"].astype(str).str.strip() == part_no.strip()
    matches = pcl[mask]
    if matches.empty:
        return {}
    row = matches.iloc[0]
    sec = str(row.get("SEC", "")) if "SEC" in pcl.columns else ""
    if sec and sec not in ("nan", ""):
        sec = sec.rstrip("0").rstrip(".") if "." in sec else sec
    return {
        "sec": sec,
        "lvl1": _s(row.get("LVL1")) if "LVL1" in pcl.columns else "",
        "lvl2": _s(row.get("LVL2")) if "LVL2" in pcl.columns else "",
        "lvl3": _s(row.get("LVL3")) if "LVL3" in pcl.columns else "",
        "lvl4": _s(row.get("LVL4")) if "LVL4" in pcl.columns else "",
        "part_name": _s(row.get("PART_NAME")),
        "qty": _s(row.get("QUANTITY")),
        "cu_fa": _s(row.get("CU_FA")),
    }


def build_vmi_groups(
    vmi_qcps: pd.DataFrame,
    pcl: pd.DataFrame,
    cit_matrix: dict[str, list[dict]],
    cu_fa_map: dict[str, str],
    ln_bom_set: set[str],
) -> list[VmiGroup]:
    """从 VMI 工序数据构建 VMI 组列表.

    返回: 每个 VmiGroup 代表一个工序 × 一种颜色的完整行集合.
    """
    if vmi_qcps.empty:
        return []

    # 按 (PROCESS_NAME, VENDOR_NAME) 分组
    groups: list[VmiGroup] = []
    grouped = vmi_qcps.groupby(["PROCESS_NAME", "VENDOR_NAME"])

    for (process, vendor), grp in grouped:
        part_nos = list(grp["PART_NO"].unique())

        # 为每个 PART_NO 查询 PCL
        pcl_info = {}
        for pn in part_nos:
            info = _lookup_pcl(pn, pcl)
            if info:
                pcl_info[pn] = info

        if not pcl_info:
            continue

        # 决定主件: 第一个有 PCL 信息的零件
        main_pn = list(pcl_info.keys())[0]

        # 检查是否有颜色 (CIT Matrix)
        color_variants: list[dict] = []
        colored_parts: dict[str, list[dict]] = {}  # {part_no: [{color_instruct_no, color_mark}]}
        for pn in part_nos:
            if pn in cit_matrix:
                variants = cit_matrix[pn]
                if variants:
                    colored_parts[pn] = variants
                    # 合并所有颜色变体 (union by color_instruct_no)
                    for cv in variants:
                        if not any(c["color_instruct_no"] == cv["color_instruct_no"] for c in color_variants):
                            color_variants.append(cv)

        # 只有主件有颜色时才做颜色展开 (SITUATION 2)
        main_colors = colored_parts.get(main_pn, [])

        if main_colors:
            # SITUATION 2: 主件有颜色 → 每种颜色一组
            for cv in main_colors:
                ci = cv["color_instruct_no"]
                cm = cv["color_mark"]
                rows = _build_vmi_rows(
                    part_nos=part_nos,
                    main_pn=main_pn,
                    pcl_info=pcl_info,
                    colored_parts=colored_parts,
                    color_instruct=ci,
                    color_mark=cm,
                    vendor=vendor,
                    cu_fa_map=cu_fa_map,
                    ln_bom_set=ln_bom_set,
                )
                if rows:
                    groups.append(VmiGroup(
                        process_name=process,
                        vendor_name=vendor,
                        color_mark=cm,
                        rows=rows,
                    ))
        else:
            # SITUATION 1: 主件无颜色 → 单体 VMI 组
            rows = _build_vmi_rows(
                part_nos=part_nos,
                main_pn=main_pn,
                pcl_info=pcl_info,
                colored_parts={},
                color_instruct="",
                color_mark="",
                vendor=vendor,
                cu_fa_map=cu_fa_map,
                ln_bom_set=ln_bom_set,
            )
            if rows:
                groups.append(VmiGroup(
                    process_name=process,
                    vendor_name=vendor,
                    color_mark="",
                    rows=rows,
                ))

    return groups


def _build_vmi_rows(
    part_nos: list[str],
    main_pn: str,
    pcl_info: dict,
    colored_parts: dict[str, list[dict]],
    color_instruct: str,
    color_mark: str,
    vendor: str,
    cu_fa_map: dict[str, str],
    ln_bom_set: set[str],
) -> list[TBRow]:
    """构建一个 VMI 组的所有 TB 行 (一种颜色).

    行结构:
      第 1 行: -S1 代表行 (SUPPLIER=vendor, USE IN=F3)
      第 2-N 行: 子件行 (SUPPLIER=原供应商, USE IN=vendor)
    """
    rows = []

    # 先构建子件行
    child_rows = []
    for pn in part_nos:
        info = pcl_info.get(pn)
        if not info:
            continue

        # 确定实际的 PART_NO (考虑颜色)
        if pn in colored_parts and color_instruct:
            actual_pn = f"{pn}-{color_instruct}"
            remarks = color_mark
        else:
            actual_pn = pn
            remarks = ""

        # SUPPLIER 从 PCL CU_FA 获取
        cu_fa = info.get("cu_fa", "")
        supplier = cu_fa_map.get(cu_fa, cu_fa)

        # EX/NEW 判定
        ex_new = "EX" if actual_pn.strip() in ln_bom_set else "NEW"

        child_rows.append(TBRow(
            sec=info.get("sec", ""),
            lvl1=info.get("lvl1", ""),
            lvl2=info.get("lvl2", ""),
            lvl3=info.get("lvl3", ""),
            lvl4=info.get("lvl4", ""),
            part_no=actual_pn,
            part_name=info.get("part_name", ""),
            qty=info.get("qty", "1"),
            supplier=supplier,
            use_in="",  # 稍后填
            ex_new=ex_new,
            remarks=remarks,
        ))

    if not child_rows:
        return []

    # 第一个子件作为主件的模板 (SEC/LVL 来源于此)
    first = child_rows[0]

    # 构建 -S1 主件行
    main_pn_base = main_pn
    if main_pn in colored_parts and color_instruct:
        main_pn_base = f"{main_pn}-{color_instruct}"
    main_part_no = f"{main_pn_base}-S1"

    main_info = pcl_info.get(main_pn, {})
    main_name = _format_subassy_name(main_info.get("part_name", ""))

    # 主件的 REMARKS: 无颜色时为空, 有颜色时填颜色标记
    main_remarks = color_mark if color_mark else ""

    main_row = TBRow(
        sec=first.sec,
        lvl1=first.lvl1,
        lvl2=first.lvl2,
        lvl3=first.lvl3,
        lvl4=first.lvl4,
        part_no=main_part_no,
        part_name=main_name,
        qty="1",
        supplier=vendor,
        use_in="F3",
        ex_new="NEW",
        remarks=main_remarks,
    )

    # 设置子件的 USE IN = vendor
    for child in child_rows:
        child.use_in = vendor

    return [main_row] + child_rows


# ---------------------------------------------------------------------------
# Part 4: 主生成函数
# ---------------------------------------------------------------------------

def generate_section_c(
    qcps_sub: pd.DataFrame,
    pcl: pd.DataFrame,
    cit_matrix: dict[str, list[dict]],
    cu_fa_map: dict[str, str],
    ln_bom_set: set[str],
) -> tuple[list[VmiGroup], list[str], int]:
    """主流程: 识别 VMI → 分组 → 构建 TB 行.

    返回: (VmiGroup 列表, 日志, 总行数)
    """
    logs = []

    # 识别 VMI 工序
    vmi_qcps = identify_vmi_processes(qcps_sub)
    if vmi_qcps.empty:
        logs.append("[WARN] 未找到 VMI 工序")
        return [], logs, 0

    n_processes = vmi_qcps["PROCESS_NAME"].nunique()
    n_parts = vmi_qcps["PART_NO"].nunique()
    logs.append(f"[STEP 4] QCPS-Sub Assy 中识别到 {n_processes} 个 VMI 工序, {len(vmi_qcps)} 个零件引用")
    logs.append(f"  唯一 PART_NO: {n_parts}")

    # 显示各工序
    for proc in sorted(vmi_qcps["PROCESS_NAME"].unique()):
        sub = vmi_qcps[vmi_qcps["PROCESS_NAME"] == proc]
        vendor = sub["VENDOR_NAME"].iloc[0]
        station = sub["STATION"].iloc[0]
        logs.append(f"  {proc}: STATION={station}, VENDOR={vendor}, 零件数={len(sub)}")

    # 构建 VMI 组
    logs.append("")
    groups = build_vmi_groups(vmi_qcps, pcl, cit_matrix, cu_fa_map, ln_bom_set)

    total_rows = 0
    for g in groups:
        n = len(g.rows)
        total_rows += n
        color_info = f" [颜色={g.color_mark}]" if g.color_mark else ""
        logs.append(
            f"[STEP 5] {g.process_name} → {g.vendor_name}{color_info}: "
            f"{n} 行 (1 -S1 + {n - 1} 子件)"
        )

    logs.append(f"\n总计: {len(groups)} 个 VMI 组, {total_rows} 行")

    return groups, logs, total_rows


# ---------------------------------------------------------------------------
# Part 5: Excel Writer
# ---------------------------------------------------------------------------

COL_HEADERS = [
    "No.", "SEC", "LVL1", "LVL2", "LVL3", "LVL4",
    "PART NO.", "PART NAME", "QTY/ BIKE", "SUPPLIER", "USE IN", "EX/NEW", "REMARKS",
]
COL_WIDTHS = [6, 8, 8, 8, 8, 8, 22, 28, 10, 18, 14, 10, 18]

THIN = Side(style="thin", color="000000")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
FONT_TITLE = Font(name="Arial", size=14, bold=True, color="FFFFFF")
FONT_HEADER = Font(name="Arial", size=11, bold=True)
FONT_SECTION = Font(name="Arial", size=11, bold=True, color="FFFFFF")
FONT_GROUP = Font(name="Arial", size=10, bold=True, color="1F4E78")
FONT_COL = Font(name="Arial", size=10, bold=True)
FONT_DATA = Font(name="Arial", size=10)
FILL_TITLE = PatternFill("solid", fgColor="1F4E78")
FILL_SECTION = PatternFill("solid", fgColor="2E75B6")
FILL_GROUP = PatternFill("solid", fgColor="E2EFDA")
FILL_COL = PatternFill("solid", fgColor="D9E1F2")
FILL_META = PatternFill("solid", fgColor="F2F2F2")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def write_section_c_excel(
    groups: list[VmiGroup],
    tb_no: str,
    target_factory: str,
    model: str,
    model_name: str,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = f"{tb_no}-SectionC"[:31]
    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    r = 1

    # 标题
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=13)
    cell = ws.cell(r, 1, value="HLYM TECHNICAL BULLETIN")
    cell.font = FONT_TITLE; cell.fill = FILL_TITLE; cell.alignment = CENTER
    ws.row_dimensions[r].height = 28
    r += 1

    # 元信息
    meta = [
        ("TB NO.", f"{tb_no} (REV. 0)"),
        ("MODEL", f"{model} - {model_name}"),
        ("PURPOSE", f"{model} ({model_name}) - VMI MASTER PART LIST FOR FACTORY {target_factory.lstrip('F')}"),
        ("WITH EFFECTIVE FROM", "TBD"),
    ]
    for label, value in meta:
        ws.cell(r, 1, value=label).font = FONT_HEADER
        ws.cell(r, 1).fill = FILL_META
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=13)
        ws.cell(r, 2, value=value).alignment = LEFT
        r += 1

    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=13)
    ws.cell(r, 1, value="The details are as listed below:").font = FONT_HEADER
    r += 2

    # Section 标题
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=13)
    sec_cell = ws.cell(r, 1, value="C. VMI PARTS - VENDOR MANAGEMENT INVENTORY")
    sec_cell.font = FONT_SECTION; sec_cell.fill = FILL_SECTION; sec_cell.alignment = LEFT
    ws.row_dimensions[r].height = 22
    r += 1

    # 列标题
    for c, h in enumerate(COL_HEADERS, 1):
        cell = ws.cell(r, c, value=h)
        cell.font = FONT_COL; cell.fill = FILL_COL; cell.alignment = CENTER; cell.border = BORDER
    r += 1

    # 按 VMI 组输出
    global_no = 0
    for grp in groups:
        # 组标题
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=13)
        color_tag = f"  |  COLOR: {grp.color_mark}" if grp.color_mark else ""
        grp_label = f"VMI: {grp.process_name}  →  Vendor: {grp.vendor_name}{color_tag}"
        grp_cell = ws.cell(r, 1, value=grp_label)
        grp_cell.font = FONT_GROUP; grp_cell.fill = FILL_GROUP; grp_cell.alignment = LEFT
        ws.row_dimensions[r].height = 20
        r += 1

        for row in grp.rows:
            global_no += 1
            vals = [
                global_no, row.sec, row.lvl1, row.lvl2, row.lvl3, row.lvl4,
                row.part_no, row.part_name, row.qty,
                row.supplier, row.use_in, row.ex_new, row.remarks,
            ]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(r, c, value=v)
                cell.font = FONT_DATA
                cell.alignment = LEFT if c in (8, 10, 11, 13) else CENTER
                cell.border = BORDER
            r += 1

        r += 1  # 组间空行

    ws.freeze_panes = "A9"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ---------------------------------------------------------------------------
# Part 6: Streamlit UI
# ---------------------------------------------------------------------------

st.set_page_config(
    page_title="TB for LCP (Section C) - VMI",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.title("TB for LCP — Section C VMI Auto Generator")
st.caption(
    "严格按 SOP 文档 'TB for LCP' 实现: "
    "QCPS-Sub Assy 识别 VMI 工序 → 查找 PCL → SITUATION 1 (无颜色) / 2 (CIT Matrix 色号) → 构建 -S1 VMI 行"
)

# ---------------------------------------------------------------------------
# 侧栏
# ---------------------------------------------------------------------------

st.sidebar.header("Data Source")
src_mode = st.sidebar.radio("Source", ["Built-in sample data", "Upload custom Excel"], index=0)
struct_arg: Optional[Path] = None

if src_mode == "Built-in sample data":
    struct_arg = STRUCT_PATH
    if struct_arg.exists():
        st.sidebar.success(f"Using: {struct_arg.name}")
    else:
        st.sidebar.error(f"File not found:\n{struct_arg}")
        st.stop()
else:
    upload_struct = st.sidebar.file_uploader(
        "TB Structure Explanation.xlsx", type=["xlsx"], key="struct"
    )
    if upload_struct:
        tmp_dir = Path(tempfile.mkdtemp())
        struct_arg = tmp_dir / upload_struct.name
        struct_arg.write_bytes(upload_struct.read())
        st.sidebar.success("File uploaded")
    else:
        st.sidebar.info("Please upload the Excel file")
        st.stop()

st.sidebar.divider()
st.sidebar.header("TB Parameters")
tb_no = st.sidebar.text_input("TB No.", value="TB-25-137")
target_factory = st.sidebar.selectbox("Target Factory", ["F1", "F2", "F3", "F4", "F5"], index=2)

run_btn = st.sidebar.button("Generate Section C", type="primary", use_container_width=True)

if not run_btn:
    st.info("Select data source on the left, then click **Generate Section C**.")
    st.stop()

# ---------------------------------------------------------------------------
# 主区: 加载数据
# ---------------------------------------------------------------------------

with st.status("Loading source data...", expanded=True) as status:
    try:
        pcl = load_pcl(struct_arg)
        cu_fa_map = load_cu_fa_code(struct_arg)
        ln_bom_set = load_ln_bom(struct_arg)
        qcps_sub = load_qcps_sub(struct_arg)
        cit_matrix = load_cit_matrix(struct_arg)
        email_map = load_email(struct_arg)
    except Exception as e:
        status.update(label=f"Load failed: {e}", state="error")
        st.exception(e)
        st.stop()

    status.write(f"PCL: **{len(pcl)}** rows")
    status.write(f"QCPS-Sub Assy: **{len(qcps_sub)}** 条 PART_NO 记录")
    status.write(f"CIT Matrix: **{len(cit_matrix)}** 有色零件")
    status.write(f"CU_FA CODE: {len(cu_fa_map)} mappings")
    status.write(f"LN BOM: {len(ln_bom_set)} parts")

    # 预览 VMI 工序
    vmi_preview = identify_vmi_processes(qcps_sub)
    if not vmi_preview.empty:
        status.write(f"VMI 工序: **{vmi_preview['PROCESS_NAME'].nunique()}** 个")
        for proc in sorted(vmi_preview["PROCESS_NAME"].unique()):
            sub = vmi_preview[vmi_preview["PROCESS_NAME"] == proc]
            status.write(f"  • {proc} → {sub['VENDOR_NAME'].iloc[0]} ({len(sub)} 零件, STATION={sub['STATION'].iloc[0]})")

    model = str(pcl.get("MODEL", pd.Series([""])).iloc[0]) if not pcl.empty else "?"
    model_name = str(pcl.get("MODEL_NAME", pd.Series([""])).iloc[0]) if not pcl.empty and "MODEL_NAME" in pcl.columns else ""
    status.write(f"Model: **{model}** ({model_name})")
    status.update(label="Data loaded", state="complete")

# ---------------------------------------------------------------------------
# 主区: 生成
# ---------------------------------------------------------------------------

with st.status("Building VMI groups...", expanded=True) as status:
    groups, logs, total_rows = generate_section_c(
        qcps_sub, pcl, cit_matrix, cu_fa_map, ln_bom_set,
    )
    for log in logs:
        status.write(log)
    status.update(label=f"VMI groups built: {len(groups)} groups, {total_rows} rows", state="complete")

# ---------------------------------------------------------------------------
# 指标
# ---------------------------------------------------------------------------

cols = st.columns(4)
cols[0].metric("Total VMI Groups", len(groups))
cols[1].metric("Total Rows", total_rows)
cols[2].metric("With Color", sum(1 for g in groups if g.color_mark))
cols[3].metric("Without Color", sum(1 for g in groups if not g.color_mark))

# ---------------------------------------------------------------------------
# 日志
# ---------------------------------------------------------------------------

with st.expander("Generation Details", expanded=False):
    for log in logs:
        if "[WARN]" in log:
            st.warning(log)
        elif "[ERROR]" in log:
            st.error(log)
        else:
            st.write(log)

# ---------------------------------------------------------------------------
# CIT Matrix 预览
# ---------------------------------------------------------------------------

with st.expander(f"CIT Matrix Preview ({len(cit_matrix)} parts)", expanded=False):
    if cit_matrix:
        preview_data = [
            {"Part No.": pn, "Color Variants": str(variants)}
            for pn, variants in sorted(cit_matrix.items())[:30]
        ]
        st.dataframe(pd.DataFrame(preview_data), use_container_width=True, hide_index=True)

# ---------------------------------------------------------------------------
# 结果表
# ---------------------------------------------------------------------------

st.subheader("Section C: VMI Parts")

if groups:
    for grp in groups:
        color_tag = f" — COLOR: **{grp.color_mark}**" if grp.color_mark else ""
        st.markdown(f"#### {grp.process_name} → Vendor: **{grp.vendor_name}**{color_tag}")

        rows_data = [
            {
                "No.": i,
                "SEC": r.sec, "LVL1": r.lvl1, "LVL2": r.lvl2, "LVL3": r.lvl3, "LVL4": r.lvl4,
                "PART NO.": r.part_no, "PART NAME": r.part_name, "QTY/BIKE": r.qty,
                "SUPPLIER": r.supplier, "USE IN": r.use_in, "EX/NEW": r.ex_new, "REMARKS": r.remarks,
            }
            for i, r in enumerate(grp.rows, 1)
        ]
        st.dataframe(pd.DataFrame(rows_data), use_container_width=True, hide_index=True)

        # 连接线提示
        st.caption(f"↳ -S1 主件 (第1行) → USE IN = F3 | 子件 (第2-{len(grp.rows)}行) → USE IN = {grp.vendor_name}")
else:
    st.warning("No VMI groups generated.")

# ---------------------------------------------------------------------------
# 下载
# ---------------------------------------------------------------------------

st.divider()
st.subheader("Download")

xlsx_bytes = write_section_c_excel(groups, tb_no, target_factory, model, model_name)

st.download_button(
    label="Download Section_C.xlsx",
    data=xlsx_bytes,
    file_name=f"{tb_no}-{target_factory}-SectionC.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    type="primary",
    use_container_width=True,
)
