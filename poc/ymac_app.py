"""TB for YMAC (Section B) Generator - Standalone Streamlit App.

严格按 SOP 文档 "TB for YMAC" (414行) 实现的 YMAC 零件 TB 自动生成系统.

SOP 情况:
  SITUATION 1 (R1-R58):   完整 PART SET NO → SOP 匹配 + "SUB ASSY."
  SITUATION 2 (R60-R242): 不完整 PART SET NO (-**) → YPL CT_COMMENT 展开 + 色型判定
  SITUATION 3 (R244-R414): 单个 YMAC 塑料件 → CIT 色号展开

SOP 步骤:
  STEP 1: 过滤 PCL.MPL_OUT_SIGN != "N"
  STEP 2: 过滤 PCL.ENGINE_SIGN != "E"
  STEP 3: 按 CU_FA 过滤 YMAC 零件 (CU_FA in {A3, A4, A5, A7})
  STEP 4: 抽取基本信息 (SITUATION 1/2/3)
  STEP 5: SUPPLIER 映射 (CU_FA CODE 表)
  STEP 6: EX/NEW 判定 (LN BOM)
  STEP 7: USE IN 判定 (QCPS 级联查询) + 最终 SUPPLIER 格式化

启动:
    streamlit run ymac_app.py
"""
from __future__ import annotations

import io
import re
import tempfile
from dataclasses import dataclass
from pathlib import Path
from typing import Optional

import pandas as pd
import streamlit as st
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parent.parent
STRUCT_PATH = PROJECT_ROOT / "NMC Project" / "[POC] TB Structure Explanation.xlsx"

YMAC_CU_FA_CODES = {"A3", "A4", "A5", "A7"}

# ---------------------------------------------------------------------------
# 行数据结构
# ---------------------------------------------------------------------------

@dataclass
class TBRow:
    sec: str = ""
    lvl1: str = ""
    lvl2: str = ""
    lvl3: str = ""
    lvl4: str = ""
    part_no: str = ""
    part_name: str = ""
    qty: str = ""
    supplier: str = ""
    use_in: str = ""
    ex_new: str = ""
    remarks: str = ""
    cu_fa: str = ""   # 内部: 用于 SUPPLIER 映射


# ---------------------------------------------------------------------------
# 工具函数
# ---------------------------------------------------------------------------

def _s(v) -> str:
    """清洗单元格值."""
    if v is None:
        return ""
    if isinstance(v, float) and pd.isna(v):
        return ""
    s = str(v).strip()
    if s.lower() in ("none", "nan"):
        return ""
    return s


# ---------------------------------------------------------------------------
# Part 1: 数据加载
# ---------------------------------------------------------------------------

def _read_sheet(path: Path, sheet_name: str, header_row: int = 0) -> pd.DataFrame:
    """读取 Excel sheet 并清洗列名、去除全空行."""
    df = pd.read_excel(path, sheet_name=sheet_name, header=header_row, engine="openpyxl")
    df.columns = [str(c).strip() if not pd.isna(c) else c for c in df.columns]
    df = df.dropna(how="all").reset_index(drop=True)
    return df


def _find_sheet(wb, keyword: str) -> str | None:
    """按关键词查找 sheet 名称."""
    kw = keyword.lower()
    for sn in wb.sheetnames:
        if kw in sn.lower():
            return sn
    return None


def load_pcl(struct_path: Path) -> pd.DataFrame:
    """加载 PCL (1. PCL)."""
    df = _read_sheet(struct_path, "1. PCL")
    rename = {
        "PARTS SET NO1": "PARTS_SET_NO",
        "PARTS_SET_NO1": "PARTS_SET_NO",
        "PART NO.": "PART_NO",
        "PART NAME": "PART_NAME",
        "QTY": "QUANTITY",
    }
    df = df.rename(columns={k: v for k, v in rename.items() if k in df.columns})
    if "PART_NO" in df.columns:
        df["PART_NO"] = df["PART_NO"].astype(str).str.strip()
        df = df[df["PART_NO"].notna() & (df["PART_NO"] != "") & (df["PART_NO"] != "nan")].reset_index(drop=True)
    if "CU_FA" in df.columns:
        df["CU_FA"] = df["CU_FA"].astype(str).str.strip()
    if "QUANTITY" in df.columns:
        df["QUANTITY"] = df["QUANTITY"].astype(str).str.strip()
    return df


def load_cu_fa_code(struct_path: Path) -> dict[str, str]:
    """加载 CU_FA CODE (6. CU_FA CODE), 返回 {code: short_name}."""
    try:
        df = _read_sheet(struct_path, "6. CU_FA CODE")
        rename = {"Customer Code": "CU_FA_CODE", "G-SU name (FROM)": "SUPPLIER_FROM"}
        df = df.rename(columns={k: v for k, v in rename.items() if k in df.columns})
        if "CU_FA_CODE" not in df.columns or "SUPPLIER_FROM" not in df.columns:
            return {}
        short_name = {
            "YMC": "YMC", "VIETNAM(YMVN)": "YMVN", "IYM(SJP)": "IYM",
            "YIMM(JKT)": "YIMM", "YIMM(WJ)": "YIMM", "TYM B/D-ASSY": "TYM",
            "YMCN": "YMCN", "TWN(YMTT) C BODY": "YMTT", "H.Y.M.M.": "HYMM",
            "H.L.Y.M.": "HLYM", "YPMV": "YPMV", "YMAC": "YMAC",
        }
        result = {}
        for _, row in df.dropna(subset=["CU_FA_CODE"]).drop_duplicates(subset=["CU_FA_CODE"]).iterrows():
            code = str(row["CU_FA_CODE"]).strip()
            full = str(row["SUPPLIER_FROM"]).strip()
            result[code] = short_name.get(full, full)
        return result
    except Exception:
        return {}


def load_ln_bom(struct_path: Path) -> set[str]:
    """加载 LN BOM (7. LN BOM), 返回 PART_NO 集合."""
    try:
        df = _read_sheet(struct_path, "7. LN BOM")
        candidates = [c for c in df.columns if str(c).strip().lower() in ("item", "part_no", "part no.")]
        col = candidates[0] if candidates else df.columns[2]
        parts = set()
        for v in df[col].astype(str).str.strip():
            if v and len(v) > 3 and v.lower() != "nan":
                parts.add(v)
        return parts
    except Exception:
        return set()


def load_qcps_sub(struct_path: Path) -> pd.DataFrame:
    """加载 QCPS Sub Assy, 处理合并单元格, 返回 PART_NO 级联索引."""
    wb = load_workbook(struct_path, read_only=True)
    sheet_name = _find_sheet(wb, "sub assy")
    wb.close()
    if not sheet_name:
        return pd.DataFrame(columns=["STATION", "PROCESS_NAME", "PART_NO"])

    wb = load_workbook(struct_path, data_only=True)
    ws = wb[sheet_name]
    max_r, max_c = ws.max_row, ws.max_column
    grid = [[ws.cell(r, c).value for c in range(1, max_c + 1)] for r in range(1, max_r + 1)]
    for rng in ws.merged_cells.ranges:
        anchor = grid[rng.min_row - 1][rng.min_col - 1]
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                grid[r - 1][c - 1] = anchor
    wb.close()

    records = []
    current_station = ""
    current_process = ""
    in_parts = False
    part_col = None

    for row in grid:
        cells_upper = [str(v).upper().strip() if v is not None else "" for v in row]
        if "STATION" in cells_upper and "PROCESS NAME" in cells_upper:
            in_parts = False
            st_idx = cells_upper.index("STATION")
            pn_idx = cells_upper.index("PROCESS NAME")
            ri = grid.index(row)
            if ri + 1 < len(grid):
                nxt = grid[ri + 1]
                current_station = _s(nxt[st_idx]).upper()
                current_process = _s(nxt[pn_idx])
            continue
        if "PART NO" in cells_upper:
            part_col = cells_upper.index("PART NO")
            in_parts = True
            continue
        if in_parts and part_col is not None:
            pn = _s(row[part_col])
            if pn:
                records.append({"STATION": current_station, "PROCESS_NAME": current_process, "PART_NO": pn})

    return pd.DataFrame(records, columns=["STATION", "PROCESS_NAME", "PART_NO"])


def load_qcps_ml(struct_path: Path, sheet_name: str) -> set[str]:
    """加载 QCPS ML(R) / ML(L), 返回 PART_NO 集合."""
    try:
        wb = load_workbook(struct_path, data_only=True)
        ws = wb[sheet_name]
        max_r, max_c = ws.max_row, ws.max_column
        grid = [[ws.cell(r, c).value for c in range(1, max_c + 1)] for r in range(1, max_r + 1)]
        for rng in ws.merged_cells.ranges:
            anchor = grid[rng.min_row - 1][rng.min_col - 1]
            for r in range(rng.min_row, rng.max_row + 1):
                for c in range(rng.min_col, rng.max_col + 1):
                    grid[r - 1][c - 1] = anchor
        wb.close()
        parts = set()
        in_parts = False
        part_col = None
        for row in grid:
            cells_upper = [str(v).upper().strip() if v is not None else "" for v in row]
            if "PART NO" in cells_upper:
                part_col = cells_upper.index("PART NO")
                in_parts = True
                continue
            if in_parts and part_col is not None:
                pn = _s(row[part_col])
                if pn:
                    parts.add(pn)
        return parts
    except Exception:
        return set()


def load_ypl(struct_path: Path) -> pd.DataFrame:
    """加载 YPL (2. YPL)."""
    for sn in ["2. YPL", "YPL"]:
        try:
            df = _read_sheet(struct_path, sn)
            rename = {
                "Part no.": "PART_NO",
                "Manufacturingassy": "MANUFACTURING_ASSY",
                "CT comment": "CT_COMMENT",
            }
            df = df.rename(columns={k: v for k, v in rename.items() if k in df.columns})
            if "PART_NO" in df.columns:
                df = df[df["PART_NO"].notna() & (df["PART_NO"] != "Part no.")].reset_index(drop=True)
            return df
        except Exception:
            continue
    return pd.DataFrame()


def load_cit(struct_path: Path) -> tuple[dict[str, list[dict]], dict[str, list[dict]]]:
    """加载 CIT TYPE A 和 TYPE B.
    返回 ({part_no: [{color_instruct_no, color_mark}, ...]}, ...)
    """
    cit_a = _load_cit_type(struct_path, "8. CIT - TYPE A")
    cit_b = _load_cit_type(struct_path, "8. CIT - TYPE B")
    return cit_a, cit_b


def _load_cit_type(struct_path: Path, sheet_name: str) -> dict[str, list[dict]]:
    """加载单个 CIT sheet, 按 part_no 索引."""
    try:
        df = _read_sheet(struct_path, sheet_name)
        part_col = None
        instruct_col = None
        mark_col = None
        for c in df.columns:
            cs = str(c).strip().lower()
            if cs in ("part no.", "part no"):
                part_col = c
            elif "color instruct" in cs:
                instruct_col = c
            elif cs == "color mark":
                mark_col = c

        if part_col is None or instruct_col is None:
            return {}

        result: dict[str, list[dict]] = {}
        for _, row in df.iterrows():
            pn = _s(row.get(part_col))
            if not pn:
                continue
            ci = _s(row.get(instruct_col))
            if not ci:
                continue
            cm = _s(row.get(mark_col)) if mark_col else ""
            if pn not in result:
                result[pn] = []
            if not any(e["color_instruct_no"] == ci for e in result[pn]):
                result[pn].append({"color_instruct_no": ci, "color_mark": cm})

        return result
    except Exception:
        return {}


def load_email(_struct_path: Path) -> dict[str, str]:
    """Structure Explanation 不含 EMAIL sheet, 返回空."""
    return {}


# ---------------------------------------------------------------------------
# Part 2: SOP Step Functions
# ---------------------------------------------------------------------------

def step1_filter_mpl(df: pd.DataFrame) -> tuple[pd.DataFrame, list[str]]:
    """STEP 1: 过滤 MPL_OUT_SIGN != 'N'."""
    logs = []
    before = len(df)
    if "MPL_OUT_SIGN" not in df.columns:
        logs.append("[SKIP] MPL_OUT_SIGN 列不存在, 跳过 STEP 1")
        return df, logs
    col = df["MPL_OUT_SIGN"].astype(str).str.strip().str.upper()
    df = df[col != "N"].reset_index(drop=True)
    after = len(df)
    logs.append(f"[STEP 1] MPL_OUT_SIGN 过滤: {before} → {after} 行 (移除 {before - after} 行)")
    return df, logs


def step2_filter_engine(df: pd.DataFrame) -> tuple[pd.DataFrame, list[str]]:
    """STEP 2: 过滤 ENGINE_SIGN != 'E'."""
    logs = []
    before = len(df)
    if "ENGINE_SIGN" not in df.columns:
        logs.append("[SKIP] ENGINE_SIGN 列不存在, 跳过 STEP 2")
        return df, logs
    col = df["ENGINE_SIGN"].astype(str).str.strip().str.upper()
    df = df[col != "E"].reset_index(drop=True)
    after = len(df)
    logs.append(f"[STEP 2] ENGINE_SIGN 过滤: {before} → {after} 行 (移除 {before - after} 行)")
    return df, logs


def step3_filter_ymac(df: pd.DataFrame) -> tuple[pd.DataFrame, list[str]]:
    """STEP 3: 过滤 CU_FA in {A3, A4, A5, A7}, 按 CU_FA 排序."""
    logs = []
    before = len(df)
    if "CU_FA" not in df.columns:
        logs.append("[ERROR] CU_FA 列不存在, 无法过滤 YMAC 零件")
        return pd.DataFrame(), logs
    mask = df["CU_FA"].astype(str).str.strip().str.upper().isin(YMAC_CU_FA_CODES)
    df = df[mask].sort_values("CU_FA").reset_index(drop=True)
    after = len(df)
    cu_counts = df["CU_FA"].value_counts().to_dict()
    logs.append(f"[STEP 3] CU_FA={{{','.join(sorted(YMAC_CU_FA_CODES))}}} 过滤: {before} → {after} 行")
    for code in sorted(cu_counts.keys()):
        logs.append(f"  CU_FA={code}: {cu_counts[code]} 行")
    return df, logs


# ---------------------------------------------------------------------------
# 辅助: 解析 CT_COMMENT / 展开 -** / 命名格式化
# ---------------------------------------------------------------------------

def _parse_ct_comment(comment: str) -> list[tuple[str, str]]:
    """解析 YPL CT_COMMENT, 如 'BBP-XH355-H0-PH(010A),BBP-XH355-J0-PJ(010B)'."""
    out = []
    for m in re.finditer(r"([A-Z0-9\-]+)\(([^)]+)\)", comment):
        out.append((m.group(1).strip(), m.group(2).strip()))
    return out


def _expand_set_no(ps: str, ypl: pd.DataFrame) -> list[tuple[str, str]]:
    """展开含 -** 的 PART SET NO 为完整编号."""

    if ypl.empty or "MANUFACTURING_ASSY" not in ypl.columns or "CT_COMMENT" not in ypl.columns:
        return []

    mask = ypl["MANUFACTURING_ASSY"].astype(str).str.strip() == ps.strip()
    matches = ypl[mask]

    for comment in matches["CT_COMMENT"].dropna().unique():
        result = _parse_ct_comment(str(comment))
        if result:
            return result

    return []


def _format_subassy_name(name: str) -> str:
    """Section B 命名: 添加 'SUB ASSY.' 后缀."""
    if not name:
        return ""
    name = name.strip()
    upper = name.upper()
    if "SUB ASSY" in upper:
        return name
    if "ASSY" in upper:
        idx = upper.index("ASSY")
        return name[:idx] + "SUB " + name[idx:]
    return f"{name} SUB ASSY."


def _tag_color_remarks(tag: str) -> str:
    """从 CT_COMMENT 标签推断 REMARKS 色型."""
    if tag:
        last = tag[-1].upper()
        if last == "A":
            return "DGNM3"
        elif last == "B":
            return "VPBC5"
    return ""


# ---------------------------------------------------------------------------
# STEP 4: 抽取基本信息 + SITUATION 1/2/3
# ---------------------------------------------------------------------------

def step4_extract_info(
    df: pd.DataFrame,
    ypl: pd.DataFrame,
    cit_a: dict[str, list[dict]],
    cit_b: dict[str, list[dict]],
) -> tuple[list[TBRow], list[str]]:
    """STEP 4: PCL → TBRow 列表, 处理 3 种 SITUATION."""
    logs = []
    rows: list[TBRow] = []

    # --- 分组: SET 组 vs 单件 ---
    has_ps = "PARTS_SET_NO" in df.columns
    ps_groups: dict[str, list[int]] = {}
    child_indices: set[int] = set()

    if has_ps:
        for idx, prow in df.iterrows():
            ps = _s(prow.get("PARTS_SET_NO"))
            if ps:
                child_indices.add(idx)
                ps_groups.setdefault(ps, []).append(idx)

    sit1_count = 0
    sit2_count = 0
    sit3_count = 0
    regular_count = 0
    warnings: list[str] = []

    # =========================================================================
    # SITUATION 1 / 2: SET 组处理
    # =========================================================================

    for ps, child_idxs in ps_groups.items():
        # SOP 匹配: PS_NO[5:9] == CHILD_PART_NO[4:8]
        anchor_idx = child_idxs[0]
        match_idx = None
        for ci in child_idxs:
            pn = _s(df.loc[ci].get("PART_NO"))
            ps_key = ps[5:9] if len(ps) >= 9 else ""
            pn_key = pn[4:8] if len(pn) >= 8 else ""
            if ps_key and pn_key and ps_key == pn_key:
                match_idx = ci
                break

        first = df.loc[anchor_idx]
        rep = df.loc[match_idx] if match_idx is not None else first

        sec = _s(first.get("SEC"))
        lvl1 = _s(first.get("LVL1"))
        lvl2 = _s(first.get("LVL2"))
        lvl3 = _s(first.get("LVL3"))
        lvl4 = _s(first.get("LVL4"))
        part_name_raw = _s(rep.get("PART_NAME"))
        cu_fa = _s(first.get("CU_FA"))

        if "**" in ps:
            # SITUATION 2: -** 展开
            expanded = _expand_set_no(ps, ypl)
            if not expanded:
                expanded = [(ps.replace("**", "00"), "")]
                warnings.append(f"PART SET NO {ps} 未在 YPL 找到, 默认填 -00")
            for exp_pn, tag in expanded:
                color_remarks = _tag_color_remarks(tag)
                pname = _format_subassy_name(part_name_raw)
                rows.append(TBRow(
                    sec=sec, lvl1=lvl1, lvl2=lvl2, lvl3=lvl3, lvl4=lvl4,
                    part_no=exp_pn, part_name=pname, qty="1",
                    remarks=color_remarks, cu_fa=cu_fa,
                ))
                sit2_count += 1
        else:
            # SITUATION 1: 完整 SET 号
            pname = _format_subassy_name(part_name_raw)
            rows.append(TBRow(
                sec=sec, lvl1=lvl1, lvl2=lvl2, lvl3=lvl3, lvl4=lvl4,
                part_no=ps, part_name=pname, qty="1", cu_fa=cu_fa,
            ))
            sit1_count += 1

    # =========================================================================
    # SITUATION 3 / 普通行: 单件处理
    # =========================================================================

    seen: set[tuple] = set()
    for idx, prow in df.iterrows():
        if idx in child_indices:
            continue

        part_no = _s(prow.get("PART_NO"))
        if not part_no:
            continue

        sec = _s(prow.get("SEC"))
        lvl1 = _s(prow.get("LVL1"))
        lvl2 = _s(prow.get("LVL2"))
        lvl3 = _s(prow.get("LVL3"))
        lvl4 = _s(prow.get("LVL4"))
        cu_fa = _s(prow.get("CU_FA"))
        part_name = _s(prow.get("PART_NAME"))
        qty = _s(prow.get("QUANTITY"))

        # 按 (SEC, LVL1, PART_NO) 去重
        key = (sec, lvl1, part_no)
        if key in seen:
            continue
        seen.add(key)

        a_entries = cit_a.get(part_no, [])
        b_entries = cit_b.get(part_no, [])

        if a_entries or b_entries:
            # SITUATION 3: CIT 色号展开
            for entry in a_entries:
                ci = entry["color_instruct_no"]
                cm = entry.get("color_mark") or "DGNM3"
                rows.append(TBRow(
                    sec=sec, lvl1=lvl1, lvl2=lvl2, lvl3=lvl3, lvl4=lvl4,
                    part_no=f"{part_no}-{ci}", part_name=part_name, qty=qty,
                    remarks=cm, cu_fa=cu_fa,
                ))
                sit3_count += 1
            for entry in b_entries:
                ci = entry["color_instruct_no"]
                cm = entry.get("color_mark") or "VPBC5"
                rows.append(TBRow(
                    sec=sec, lvl1=lvl1, lvl2=lvl2, lvl3=lvl3, lvl4=lvl4,
                    part_no=f"{part_no}-{ci}", part_name=part_name, qty=qty,
                    remarks=cm, cu_fa=cu_fa,
                ))
                sit3_count += 1
        else:
            # 普通行 (无色号展开)
            rows.append(TBRow(
                sec=sec, lvl1=lvl1, lvl2=lvl2, lvl3=lvl3, lvl4=lvl4,
                part_no=part_no, part_name=part_name, qty=qty, cu_fa=cu_fa,
            ))
            regular_count += 1

    logs.append(f"[STEP 4] 抽取完成: {len(rows)} 行")
    logs.append(f"  SITUATION 1 (完整 SET): {sit1_count} 行")
    logs.append(f"  SITUATION 2 (-** 展开): {sit2_count} 行")
    logs.append(f"  SITUATION 3 (CIT 色号展开): {sit3_count} 行")
    logs.append(f"  普通行: {regular_count} 行")
    for w in warnings:
        logs.append(f"  [WARN] {w}")

    return rows, logs


# ---------------------------------------------------------------------------
# STEP 5: SUPPLIER 映射
# ---------------------------------------------------------------------------

def step5_supplier(rows: list[TBRow], cu_fa_map: dict[str, str]) -> tuple[list[TBRow], list[str]]:
    """STEP 5: 根据 CU_FA 查 CU_FA CODE 表 → SUPPLIER."""
    logs = []
    code_counts: dict[str, int] = {}
    for row in rows:
        supplier_name = cu_fa_map.get(row.cu_fa, row.cu_fa)
        code_counts[row.cu_fa] = code_counts.get(row.cu_fa, 0) + 1
        if row.use_in and row.use_in.upper() != "F3":
            row.supplier = f"{supplier_name} TO F3"
        else:
            row.supplier = supplier_name

    for code, cnt in sorted(code_counts.items()):
        name = cu_fa_map.get(code, code)
        logs.append(f"  CU_FA={code} → {name} ({cnt} 行)")
    logs.append(f"[STEP 5] SUPPLIER 映射完成: {len(code_counts)} 个供应商代码")
    return rows, logs


# ---------------------------------------------------------------------------
# STEP 6: EX/NEW 判定
# ---------------------------------------------------------------------------

def step6_ex_new(rows: list[TBRow], ln_bom_set: set[str]) -> tuple[list[TBRow], list[str]]:
    """STEP 6: PART_NO in LN BOM → EX, else → NEW."""
    logs = []
    ex_count = 0
    new_count = 0
    for row in rows:
        pn = row.part_no.strip()
        if pn in ln_bom_set:
            row.ex_new = "EX"
            ex_count += 1
        else:
            row.ex_new = "NEW"
            new_count += 1
    logs.append(f"[STEP 6] EX/NEW 判定: EX={ex_count}, NEW={new_count}")
    return rows, logs


# ---------------------------------------------------------------------------
# STEP 7: USE IN 判定
# ---------------------------------------------------------------------------

def step7_use_in(
    rows: list[TBRow],
    qcps_sub: pd.DataFrame,
    qcps_mlr: set[str],
    qcps_mll: set[str],
    email_map: dict[str, str],
    cu_fa_map: dict[str, str],
) -> tuple[list[TBRow], list[str]]:
    """STEP 7: USE IN 级联查询 + 最终 SUPPLIER 格式化.

    级联:
      1. QCPS Sub Assy + STATION = EXTERNAL VENDOR → email 或 MODERNRIA
      2. QCPS Sub Assy + STATION = SUB ASSY    → F3
      3. QCPS ML(R)                            → F3
      4. QCPS ML(L)                            → F3
      5. Default                               → F3
    """
    logs = []

    # 构建 Sub Assy 索引: {PART_NO: {STATION, PROCESS_NAME}}
    sub_idx: dict[str, dict] = {}
    if not qcps_sub.empty and "PART_NO" in qcps_sub.columns:
        for _, r in qcps_sub.iterrows():
            pn = str(r.get("PART_NO", "")).strip()
            if not pn:
                continue
            station = str(r.get("STATION", "")).strip().upper()
            process = str(r.get("PROCESS_NAME", "")).strip()
            if pn not in sub_idx:
                sub_idx[pn] = {"STATION": station, "PROCESS_NAME": process}
            elif "EXTERNAL VENDOR" in station and "EXTERNAL VENDOR" not in sub_idx[pn]["STATION"]:
                sub_idx[pn] = {"STATION": station, "PROCESS_NAME": process}

    f3_count = 0
    vendor_count = 0

    for row in rows:
        pn = row.part_no.strip()
        use_in = "F3"

        if pn in sub_idx:
            info = sub_idx[pn]
            if "EXTERNAL VENDOR" in info["STATION"]:
                use_in = email_map.get(pn, "MODERNRIA")
                vendor_count += 1
            else:
                f3_count += 1
        elif pn in qcps_mlr:
            f3_count += 1
        elif pn in qcps_mll:
            f3_count += 1
        else:
            f3_count += 1

        row.use_in = use_in

        # 最终 SUPPLIER 格式化: use_in != F3 → "SUPPLIER TO F3"
        supplier_name = cu_fa_map.get(row.cu_fa, row.cu_fa)
        if row.use_in and row.use_in.upper() != "F3":
            row.supplier = f"{supplier_name} TO F3"
        else:
            row.supplier = supplier_name

    logs.append(f"[STEP 7] USE IN 判定: F3={f3_count}, External Vendor={vendor_count}")
    return rows, logs


# ---------------------------------------------------------------------------
# Part 3: 主生成函数
# ---------------------------------------------------------------------------

def generate_section_b(
    pcl: pd.DataFrame,
    cu_fa_map: dict[str, str],
    ln_bom_set: set[str],
    qcps_sub: pd.DataFrame,
    qcps_mlr: set[str],
    qcps_mll: set[str],
    ypl: pd.DataFrame,
    cit_a: dict[str, list[dict]],
    cit_b: dict[str, list[dict]],
    email_map: dict[str, str],
) -> tuple[list[TBRow], list[str]]:
    """主流程: STEP 1-7 全部串联."""
    all_logs: list[str] = []

    df, logs = step1_filter_mpl(pcl)
    all_logs.extend(logs)

    df, logs = step2_filter_engine(df)
    all_logs.extend(logs)

    df, logs = step3_filter_ymac(df)
    all_logs.extend(logs)

    if df.empty:
        all_logs.append("[ERROR] 过滤后无 YMAC 数据，终止生成")
        return [], all_logs

    rows, logs = step4_extract_info(df, ypl, cit_a, cit_b)
    all_logs.extend(logs)

    rows, logs = step5_supplier(rows, cu_fa_map)
    all_logs.extend(logs)

    rows, logs = step6_ex_new(rows, ln_bom_set)
    all_logs.extend(logs)

    rows, logs = step7_use_in(rows, qcps_sub, qcps_mlr, qcps_mll, email_map, cu_fa_map)
    all_logs.extend(logs)

    return rows, all_logs


# ---------------------------------------------------------------------------
# Part 4: Excel Writer
# ---------------------------------------------------------------------------

COL_HEADERS = [
    "No.", "SEC", "LVL1", "LVL2", "LVL3", "LVL4",
    "PART NO.", "PART NAME", "QTY/ BIKE", "SUPPLIER", "USE IN", "EX/NEW", "REMARKS",
]
COL_WIDTHS = [6, 8, 8, 8, 8, 8, 22, 28, 10, 18, 14, 10, 18]

THIN = Side(style="thin", color="000000")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
FONT_TITLE = Font(name="Arial", size=14, bold=True, color="FFFFFF")
FONT_HEADER = Font(name="Arial", size=11, bold=True)
FONT_SECTION = Font(name="Arial", size=11, bold=True, color="FFFFFF")
FONT_COL = Font(name="Arial", size=10, bold=True)
FONT_DATA = Font(name="Arial", size=10)
FILL_TITLE = PatternFill("solid", fgColor="1F4E78")
FILL_SECTION = PatternFill("solid", fgColor="2E75B6")
FILL_COL = PatternFill("solid", fgColor="D9E1F2")
FILL_META = PatternFill("solid", fgColor="F2F2F2")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def write_section_b_excel(
    rows: list[TBRow],
    tb_no: str,
    target_factory: str,
    model: str,
    model_name: str,
) -> bytes:
    """将 Section B 行写入格式化 .xlsx."""
    wb = Workbook()
    ws = wb.active
    ws.title = f"{tb_no}-SectionB"[:31]

    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    r = 1

    # 标题
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=13)
    cell = ws.cell(r, 1, value="HLYM TECHNICAL BULLETIN")
    cell.font = FONT_TITLE
    cell.fill = FILL_TITLE
    cell.alignment = CENTER
    ws.row_dimensions[r].height = 28
    r += 1

    # 元信息
    meta = [
        ("TB NO.", f"{tb_no} (REV. 0)"),
        ("MODEL", f"{model} - {model_name}"),
        ("PURPOSE", f"{model} ({model_name}) - MASTER PART LIST FOR FACTORY {target_factory.lstrip('F')}"),
        ("WITH EFFECTIVE FROM", "TBD"),
    ]
    for label, value in meta:
        ws.cell(r, 1, value=label).font = FONT_HEADER
        ws.cell(r, 1).fill = FILL_META
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=13)
        ws.cell(r, 2, value=value).alignment = LEFT
        r += 1

    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=13)
    ws.cell(r, 1, value="The details are as listed below:").font = FONT_HEADER
    r += 2

    # Section 标题
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=13)
    sec_cell = ws.cell(r, 1, value="B. PARTS WHICH SUPPLY FROM YMTT, YMCN & IYM, YIMM")
    sec_cell.font = FONT_SECTION
    sec_cell.fill = FILL_SECTION
    sec_cell.alignment = LEFT
    ws.row_dimensions[r].height = 22
    r += 1

    # 列标题
    for c, h in enumerate(COL_HEADERS, 1):
        cell = ws.cell(r, c, value=h)
        cell.font = FONT_COL
        cell.fill = FILL_COL
        cell.alignment = CENTER
        cell.border = BORDER
    r += 1

    # 数据行
    for i, row in enumerate(rows, 1):
        vals = [
            i, row.sec, row.lvl1, row.lvl2, row.lvl3, row.lvl4,
            row.part_no, row.part_name, row.qty,
            row.supplier, row.use_in, row.ex_new, row.remarks,
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(r, c, value=v)
            cell.font = FONT_DATA
            cell.alignment = LEFT if c in (8, 10, 11, 13) else CENTER
            cell.border = BORDER
        r += 1

    ws.freeze_panes = "A9"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ---------------------------------------------------------------------------
# Part 5: Streamlit UI
# ---------------------------------------------------------------------------

st.set_page_config(
    page_title="TB for YMAC (Section B)",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.title("TB for YMAC — Section B Auto Generator")
st.caption(
    "严格按 SOP 文档 'TB for YMAC' 实现: "
    "STEP 1-3 过滤 → STEP 4 (SITUATION 1/2/3) → STEP 5 供应商 → STEP 6 EX/NEW → STEP 7 USE IN"
)

# ---------------------------------------------------------------------------
# 侧栏
# ---------------------------------------------------------------------------

st.sidebar.header("Data Source")

src_mode = st.sidebar.radio("Source", ["Built-in sample data", "Upload custom Excel"], index=0)

struct_arg: Optional[Path] = None

if src_mode == "Built-in sample data":
    struct_arg = STRUCT_PATH
    if struct_arg.exists():
        st.sidebar.success(f"Using: {struct_arg.name}")
    else:
        st.sidebar.error(f"File not found:\n{struct_arg}")
        st.stop()
else:
    upload_struct = st.sidebar.file_uploader(
        "TB Structure Explanation.xlsx", type=["xlsx"], key="struct"
    )
    if upload_struct:
        tmp_dir = Path(tempfile.mkdtemp())
        struct_arg = tmp_dir / upload_struct.name
        struct_arg.write_bytes(upload_struct.read())
        st.sidebar.success("File uploaded")
    else:
        st.sidebar.info("Please upload the Excel file")
        st.stop()

st.sidebar.divider()
st.sidebar.header("TB Parameters")
tb_no = st.sidebar.text_input("TB No.", value="TB-25-137")
target_factory = st.sidebar.selectbox(
    "Target Factory", ["F1", "F2", "F3", "F4", "F5"], index=2
)

run_btn = st.sidebar.button(
    "Generate Section B", type="primary", use_container_width=True
)

if not run_btn:
    st.info("Select data source on the left, then click **Generate Section B**.")
    st.stop()

# ---------------------------------------------------------------------------
# 主区: 加载数据
# ---------------------------------------------------------------------------

with st.status("Loading source data...", expanded=True) as status:
    try:
        pcl = load_pcl(struct_arg)
        cu_fa_map = load_cu_fa_code(struct_arg)
        ln_bom_set = load_ln_bom(struct_arg)
        qcps_sub = load_qcps_sub(struct_arg)
        qcps_mlr = load_qcps_ml(struct_arg, "4. QCPS - ML (R)")
        qcps_mll = load_qcps_ml(struct_arg, "5. QCPS - ML (L)")
        ypl = load_ypl(struct_arg)
        cit_a, cit_b = load_cit(struct_arg)
        email_map = load_email(struct_arg)
    except Exception as e:
        status.update(label=f"Load failed: {e}", state="error")
        st.exception(e)
        st.stop()

    status.write(f"PCL: **{len(pcl)}** rows")
    if "CU_FA" in pcl.columns:
        ymac_mask = (
            pcl["CU_FA"].astype(str).str.strip().str.upper().isin(YMAC_CU_FA_CODES)
        )
        codes_str = ",".join(sorted(YMAC_CU_FA_CODES))
        status.write(f"  YMAC CU_FA ({codes_str}): {ymac_mask.sum()} rows")
    status.write(f"CU_FA CODE: {len(cu_fa_map)} mappings")
    for code in sorted(YMAC_CU_FA_CODES):
        status.write(f"  {code} → {cu_fa_map.get(code, '?')}")
    status.write(
        f"LN BOM: {len(ln_bom_set)} parts  |  "
        f"QCPS Sub: {len(qcps_sub)}  |  ML(R): {len(qcps_mlr)}  |  ML(L): {len(qcps_mll)}"
    )
    status.write(
        f"YPL: {len(ypl)} rows  |  CIT Type A: {len(cit_a)} parts  |  CIT Type B: {len(cit_b)} parts"
    )

    model = str(pcl.get("MODEL", pd.Series([""])).iloc[0]) if not pcl.empty else "?"
    model_name_col = "MODEL_NAME" if "MODEL_NAME" in pcl.columns else None
    model_name = (
        str(pcl.get(model_name_col, pd.Series([""])).iloc[0])
        if model_name_col and not pcl.empty
        else ""
    )
    status.write(f"Model: **{model}** ({model_name})")
    status.update(label="Data loaded", state="complete")

# ---------------------------------------------------------------------------
# 主区: 生成
# ---------------------------------------------------------------------------

with st.status("Applying SOP rules...", expanded=True) as status:
    rows, logs = generate_section_b(
        pcl, cu_fa_map, ln_bom_set, qcps_sub, qcps_mlr, qcps_mll,
        ypl, cit_a, cit_b, email_map,
    )
    for log in logs:
        status.write(log)
    status.update(label=f"Section B generated: {len(rows)} rows", state="complete")

# ---------------------------------------------------------------------------
# 指标
# ---------------------------------------------------------------------------

ex_count = sum(1 for r in rows if r.ex_new == "EX")
new_count = sum(1 for r in rows if r.ex_new == "NEW")
f3_count = sum(1 for r in rows if r.use_in == "F3")
vendor_count = sum(1 for r in rows if r.use_in != "F3")

cols = st.columns(5)
cols[0].metric("Total Parts", len(rows))
cols[1].metric("EX", ex_count)
cols[2].metric("NEW", new_count)
cols[3].metric("USE IN = F3", f3_count)
cols[4].metric("USE IN = Vendor", vendor_count)

# ---------------------------------------------------------------------------
# SOP 详细日志
# ---------------------------------------------------------------------------

with st.expander("SOP Step Details", expanded=False):
    for log in logs:
        if "[ERROR]" in log:
            st.error(log)
        elif "[WARN]" in log:
            st.warning(log)
        elif "[SKIP]" in log:
            st.info(log)
        else:
            st.write(log)

# ---------------------------------------------------------------------------
# 结果表
# ---------------------------------------------------------------------------

st.subheader("Section B: PARTS WHICH SUPPLY FROM YMTT, YMCN & IYM, YIMM")

if rows:
    rows_data = [
        {
            "No.": i,
            "SEC": r.sec,
            "LVL1": r.lvl1,
            "LVL2": r.lvl2,
            "LVL3": r.lvl3,
            "LVL4": r.lvl4,
            "PART NO.": r.part_no,
            "PART NAME": r.part_name,
            "QTY/BIKE": r.qty,
            "SUPPLIER": r.supplier,
            "USE IN": r.use_in,
            "EX/NEW": r.ex_new,
            "REMARKS": r.remarks,
        }
        for i, r in enumerate(rows, 1)
    ]
    st.dataframe(pd.DataFrame(rows_data), use_container_width=True, hide_index=True)
else:
    st.warning("No Section B rows generated.")

# ---------------------------------------------------------------------------
# 下载
# ---------------------------------------------------------------------------

st.divider()
st.subheader("Download")

xlsx_bytes = write_section_b_excel(rows, tb_no, target_factory, model, model_name)

st.download_button(
    label="Download Section_B.xlsx",
    data=xlsx_bytes,
    file_name=f"{tb_no}-{target_factory}-SectionB.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    type="primary",
    use_container_width=True,
)
