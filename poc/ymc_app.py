"""TB for YMC (Section A) Generator - Standalone Streamlit App.

严格按 SOP 文档 "TB for YMC" (601行) 实现的 YMC 零件 TB 自动生成系统.

SOP 步骤:
  STEP 1: 过滤 PCL.MPL_OUT_SIGN != "N"
  STEP 2: 过滤 PCL.ENGINE_SIGN != "E"
  STEP 3: 按 CU_FA 过滤 YMC 零件 (CU_FA = "A2")
  STEP 4: 抽取基本信息 (SEC/LVL/PART_NO/PART_NAME/QTY)
  STEP 5: SUPPLIER 映射 (CU_FA CODE 表)
  STEP 6: EX/NEW 判定 (LN BOM)
  STEP 7: USE IN 判定 (QCPS 级联查询)
  SITUATION 2: 不完整 PART SET NO (-**)
  SITUATION 3: 完整 PART SET NO

启动:
    streamlit run ymc_app.py
"""
from __future__ import annotations

import re
import tempfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

import pandas as pd
import streamlit as st
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parent.parent
SAMPLE_PATH = PROJECT_ROOT / "NMC Project" / "[POC] Sample TB.xlsx"
STRUCT_PATH = PROJECT_ROOT / "NMC Project" / "[POC] TB Structure Explanation.xlsx"


# ---------------------------------------------------------------------------
# 行数据结构
# ---------------------------------------------------------------------------

@dataclass
class TBRow:
    sec: str = ""
    lvl1: str = ""
    lvl2: str = ""
    lvl3: str = ""
    lvl4: str = ""
    part_no: str = ""
    part_name: str = ""
    qty: str = ""
    supplier: str = ""
    use_in: str = ""
    ex_new: str = ""
    remarks: str = ""


# ---------------------------------------------------------------------------
# 工具函数
# ---------------------------------------------------------------------------

def _s(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and pd.isna(v):
        return ""
    s = str(v).strip()
    if s.lower() in ("none", "nan"):
        return ""
    return s


# ---------------------------------------------------------------------------
# Part 1: 数据加载
# ---------------------------------------------------------------------------

def _read_sheet(path: Path, sheet_name: str, header_row: int = 0) -> pd.DataFrame:
    df = pd.read_excel(path, sheet_name=sheet_name, header=header_row, engine="openpyxl")
    df.columns = [str(c).strip() if not pd.isna(c) else c for c in df.columns]
    df = df.dropna(how="all").reset_index(drop=True)
    return df


def _load_pcl_from(path: Path, sheet_name: str) -> pd.DataFrame:
    """从单个文件加载 PCL."""
    df = _read_sheet(path, sheet_name)
    rename = {
        "PARTS SET NO1": "PARTS_SET_NO",
        "PARTS_SET_NO1": "PARTS_SET_NO",
        "PART NO.": "PART_NO",
        "PART NAME": "PART_NAME",
        "QTY": "QUANTITY",
    }
    df = df.rename(columns={k: v for k, v in rename.items() if k in df.columns})
    if "PART_NO" in df.columns:
        df["PART_NO"] = df["PART_NO"].astype(str).str.strip()
        df = df[df["PART_NO"].notna() & (df["PART_NO"] != "") & (df["PART_NO"] != "nan")].reset_index(drop=True)
    if "CU_FA" in df.columns:
        df["CU_FA"] = df["CU_FA"].astype(str).str.strip()
    if "QUANTITY" in df.columns:
        df["QUANTITY"] = df["QUANTITY"].astype(str).str.strip()
    return df


def load_pcl(struct_path: Path) -> pd.DataFrame:
    """加载 PCL (从 Structure Explanation)."""
    return _load_pcl_from(struct_path, "1. PCL")


def load_cu_fa_code(path: Path) -> dict[str, str]:
    """加载 CU_FA CODE 表, 返回 {Customer Code: supplier_short_name}."""
    try:
        df = _read_sheet(path, "6. CU_FA CODE")
        rename = {"Customer Code": "CU_FA_CODE", "G-SU name (FROM)": "SUPPLIER_FROM"}
        df = df.rename(columns={k: v for k, v in rename.items() if k in df.columns})
        if "CU_FA_CODE" not in df.columns or "SUPPLIER_FROM" not in df.columns:
            return {}
        short_name = {
            "YMC": "YMC", "VIETNAM(YMVN)": "YMVN", "IYM(SJP)": "IYM",
            "YIMM(JKT)": "YIMM", "YIMM(WJ)": "YIMM", "TYM B/D-ASSY": "TYM",
            "YMCN": "YMCN", "TWN(YMTT) C BODY": "YMTT", "H.Y.M.M.": "HYMM",
            "H.L.Y.M.": "HLYM", "YPMV": "YPMV", "YMAC": "YMAC",
        }
        result = {}
        for _, row in df.dropna(subset=["CU_FA_CODE"]).drop_duplicates(subset=["CU_FA_CODE"]).iterrows():
            code = str(row["CU_FA_CODE"]).strip()
            full = str(row["SUPPLIER_FROM"]).strip()
            result[code] = short_name.get(full, full)
        return result
    except Exception:
        return {}


def load_ln_bom(path: Path) -> set[str]:
    """加载 LN BOM, 返回 PART_NO 集合."""
    try:
        df = _read_sheet(path, "7. LN BOM")
        candidates = [c for c in df.columns if str(c).strip().lower() in ("item", "part_no", "part no.")]
        if candidates:
            col = candidates[0]
        else:
            col = df.columns[2]
        parts = set()
        for v in df[col].astype(str).str.strip():
            if v and len(v) > 3 and v.lower() != "nan":
                parts.add(v)
        return parts
    except Exception:
        return set()


def load_qcps_sub(struct_path: Path) -> pd.DataFrame:
    """加载 QCPS Sub Assy (块结构解析)."""
    sheet_name = None
    wb = load_workbook(struct_path, read_only=True)
    for sn in wb.sheetnames:
        if "sub assy" in sn.lower() or "subassy" in sn.lower():
            sheet_name = sn
            break
    wb.close()
    if not sheet_name:
        return pd.DataFrame(columns=["STATION", "PROCESS_NAME", "PART_NO"])

    wb = load_workbook(struct_path, data_only=True)
    ws = wb[sheet_name]
    max_r, max_c = ws.max_row, ws.max_column
    grid = [[ws.cell(r, c).value for c in range(1, max_c + 1)] for r in range(1, max_r + 1)]
    for rng in ws.merged_cells.ranges:
        anchor = grid[rng.min_row - 1][rng.min_col - 1]
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                grid[r - 1][c - 1] = anchor
    wb.close()

    records = []
    current_station = ""
    current_process = ""
    in_parts = False
    part_col = None

    for row in grid:
        cells_upper = [str(v).upper().strip() if v is not None else "" for v in row]
        if "STATION" in cells_upper and "PROCESS NAME" in cells_upper:
            in_parts = False
            st_idx = cells_upper.index("STATION")
            pn_idx = cells_upper.index("PROCESS NAME")
            ri = grid.index(row)
            if ri + 1 < len(grid):
                nxt = grid[ri + 1]
                current_station = _s(nxt[st_idx]).upper()
                current_process = _s(nxt[pn_idx])
            continue
        if "PART NO" in cells_upper:
            part_col = cells_upper.index("PART NO")
            in_parts = True
            continue
        if in_parts and part_col is not None:
            pn = _s(row[part_col])
            if pn:
                records.append({"STATION": current_station, "PROCESS_NAME": current_process, "PART_NO": pn})

    return pd.DataFrame(records, columns=["STATION", "PROCESS_NAME", "PART_NO"])


def load_qcps_ml(struct_path: Path, sheet_name: str) -> set[str]:
    """加载 QCPS ML(R)/ML(L), 返回 PART_NO 集合."""
    try:
        wb = load_workbook(struct_path, data_only=True)
        ws = wb[sheet_name]
        max_r, max_c = ws.max_row, ws.max_column
        grid = [[ws.cell(r, c).value for c in range(1, max_c + 1)] for r in range(1, max_r + 1)]
        for rng in ws.merged_cells.ranges:
            anchor = grid[rng.min_row - 1][rng.min_col - 1]
            for r in range(rng.min_row, rng.max_row + 1):
                for c in range(rng.min_col, rng.max_col + 1):
                    grid[r - 1][c - 1] = anchor
        wb.close()

        parts = set()
        in_parts = False
        part_col = None
        for row in grid:
            cells_upper = [str(v).upper().strip() if v is not None else "" for v in row]
            if "PART NO" in cells_upper:
                part_col = cells_upper.index("PART NO")
                in_parts = True
                continue
            if in_parts and part_col is not None:
                pn = _s(row[part_col])
                if pn:
                    parts.add(pn)
        return parts
    except Exception:
        return set()


def load_ypl(struct_path: Path) -> pd.DataFrame:
    """加载 YPL (用于 PART SET NO -** 完整化)."""
    for sn in ["2. YPL", "YPL"]:
        try:
            df = _read_sheet(struct_path, sn)
            rename = {
                "Part no.": "PART_NO",
                "Manufacturingassy": "MANUFACTURING_ASSY",
                "CT comment": "CT_COMMENT",
            }
            df = df.rename(columns={k: v for k, v in rename.items() if k in df.columns})
            if "PART_NO" in df.columns:
                df = df[df["PART_NO"].notna() & (df["PART_NO"] != "Part no.")].reset_index(drop=True)
            return df
        except Exception:
            continue
    return pd.DataFrame()


def load_email(struct_path: Path) -> dict[str, str]:
    """加载 EMAIL 本地供应商. Structure Explanation 无此 sheet 时返回空 dict."""
    return {}


def load_ground_truth(path: Path, sheet_name: str = "TB-25-137-F3_1") -> list[dict]:
    """读取 ground truth sheet, 返回仅 Section A 的行."""
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet_name]
    rows_data = list(ws.iter_rows(values_only=True))
    wb.close()

    truth = []
    current_section = ""
    in_data = False

    for row in rows_data:
        cells = [c for c in row]
        if all(c is None or str(c).strip() == "" for c in cells):
            continue
        a = str(cells[0]).strip() if cells[0] is not None else ""
        b = str(cells[1]).strip() if len(cells) > 1 and cells[1] is not None else ""

        if a and len(a) == 2 and a.endswith(".") and a[0].isalpha():
            current_section = a[0]
            in_data = False
            continue
        if a == "No." and b == "SEC":
            in_data = True
            continue
        if in_data and cells[0] is not None:
            try:
                int(cells[0])
            except (ValueError, TypeError):
                continue
            if current_section != "A":
                continue
            def safe(idx):
                if idx >= len(cells) or cells[idx] is None:
                    return ""
                return str(cells[idx]).strip()
            truth.append({
                "section": current_section,
                "sec": safe(1), "lvl1": safe(2), "lvl2": safe(3),
                "lvl3": safe(4), "lvl4": safe(5),
                "part_no": safe(6), "part_name": safe(7), "qty": safe(8),
                "supplier": safe(9), "use_in": safe(10), "ex_new": safe(11), "remarks": safe(12),
            })
    return truth


# ---------------------------------------------------------------------------
# Part 2: SOP Step Functions
# ---------------------------------------------------------------------------

def step1_filter_mpl(df: pd.DataFrame) -> tuple[pd.DataFrame, list[str]]:
    """STEP 1: 过滤 MPL_OUT_SIGN != 'N' (保留 Y 或空)."""
    logs = []
    before = len(df)
    if "MPL_OUT_SIGN" not in df.columns:
        logs.append(f"[SKIP] MPL_OUT_SIGN 列不存在, 跳过 STEP 1 过滤")
        return df, logs
    col = df["MPL_OUT_SIGN"].astype(str).str.strip().str.upper()
    df = df[col != "N"].reset_index(drop=True)
    after = len(df)
    removed = before - after
    logs.append(f"[STEP 1] MPL_OUT_SIGN 过滤: {before} → {after} 行 (移除 {removed} 行 MPL_OUT_SIGN='N')")
    return df, logs


def step2_filter_engine(df: pd.DataFrame) -> tuple[pd.DataFrame, list[str]]:
    """STEP 2: 过滤 ENGINE_SIGN != 'E' (移除发动机件)."""
    logs = []
    before = len(df)
    if "ENGINE_SIGN" not in df.columns:
        logs.append(f"[SKIP] ENGINE_SIGN 列不存在, 跳过 STEP 2 过滤")
        return df, logs
    col = df["ENGINE_SIGN"].astype(str).str.strip().str.upper()
    df = df[col != "E"].reset_index(drop=True)
    after = len(df)
    removed = before - after
    logs.append(f"[STEP 2] ENGINE_SIGN 过滤: {before} → {after} 行 (移除 {removed} 行 ENGINE_SIGN='E')")
    return df, logs


def step3_filter_ymc(df: pd.DataFrame) -> tuple[pd.DataFrame, list[str]]:
    """STEP 3: 过滤 CU_FA = 'A2' (仅 YMC 零件), 按 CU_FA 排序."""
    logs = []
    before = len(df)
    if "CU_FA" not in df.columns:
        logs.append("[ERROR] CU_FA 列不存在, 无法过滤 YMC 零件")
        return pd.DataFrame(), logs
    df = df[df["CU_FA"].astype(str).str.strip().str.upper() == "A2"].reset_index(drop=True)
    after = len(df)
    logs.append(f"[STEP 3] CU_FA='A2' 过滤: {before} → {after} 行 (仅 YMC 供应商零件)")
    return df, logs


def _parse_ct_comment(comment: str) -> list[tuple[str, str]]:
    """解析 YPL CT_COMMENT, 如 'BBP-XH355-H0-PH(010A),BBP-XH355-J0-PJ(010B)'."""
    out = []
    for m in re.finditer(r"([A-Z0-9\-]+)\(([^)]+)\)", comment):
        out.append((m.group(1).strip(), m.group(2).strip()))
    return out


def step4_extract_info(df: pd.DataFrame, ypl: pd.DataFrame) -> tuple[list[TBRow], list[str]]:
    """STEP 4: 抽取基本信息 + 处理 PART SET NO (SITUATION 2/3).

    SOP 规则:
    - 无 PARTS_SET_NO 的行 → 直接输出 (SITUATION 1)
    - 有 PARTS_SET_NO 的行:
      - SOP 匹配: PS_NO[5:9] == PART_NO[4:8] → 选出代表行
      - 含 -** → 用 YPL CT_COMMENT 展开 (SITUATION 2)
      - 不含 -** → 直接用完整 SET 号 (SITUATION 3)
    """
    logs = []
    rows: list[TBRow] = []

    # --- 预处理: 分离 SET 子件和普通行 ---
    has_ps = "PARTS_SET_NO" in df.columns
    ps_groups: dict[str, list[int]] = {}  # PS_NO → [df indices]
    child_indices: set[int] = set()

    if has_ps:
        for idx, prow in df.iterrows():
            ps = _s(prow.get("PARTS_SET_NO"))
            if ps:
                child_indices.add(idx)
                ps_groups.setdefault(ps, []).append(idx)

    # --- 处理每个 SET 组 ---
    set_warnings = []
    for ps, child_idxs in ps_groups.items():
        # SOP 匹配: 找 PS_NO[5:9] == PART_NO[4:8] 的子件
        anchor_idx = child_idxs[0]
        match_idx = None
        for ci in child_idxs:
            pn = _s(df.loc[ci].get("PART_NO"))
            ps_key = ps[5:9] if len(ps) >= 9 else ""
            pn_key = pn[4:8] if len(pn) >= 8 else ""
            if ps_key and pn_key and ps_key == pn_key:
                match_idx = ci
                break

        # SEC/LVL 用首个子件, PART_NAME 用 SOP 匹配子件
        first = df.loc[anchor_idx]
        rep = df.loc[match_idx] if match_idx is not None else first
        sec = _s(first.get("SEC"))
        lvl1 = _s(first.get("LVL1"))
        lvl2 = _s(first.get("LVL2"))
        lvl3 = _s(first.get("LVL3"))
        lvl4 = _s(first.get("LVL4"))
        part_name_raw = _s(rep.get("PART_NAME"))

        # SITUATION 2: -** 展开
        if "**" in ps:
            expanded = _expand_set_no(ps, ypl)
            if not expanded:
                expanded = [(ps.replace("**", "00"), "")]
                set_warnings.append(f"PART SET NO {ps} 未在 YPL 找到, 默认填 -00")
            for exp_pn, color_tag in expanded:
                pname = _format_assy_name(part_name_raw)
                rows.append(TBRow(
                    sec=sec, lvl1=lvl1, lvl2=lvl2, lvl3=lvl3, lvl4=lvl4,
                    part_no=exp_pn, part_name=pname, qty="1",
                    remarks=color_tag,
                ))
        else:
            # SITUATION 3: 完整 SET 号
            pname = _format_assy_name(part_name_raw)
            rows.append(TBRow(
                sec=sec, lvl1=lvl1, lvl2=lvl2, lvl3=lvl3, lvl4=lvl4,
                part_no=ps, part_name=pname, qty="1",
            ))

    # --- 处理普通行 (非 SET 子件) ---
    seen: set[tuple] = set()
    for idx, prow in df.iterrows():
        if idx in child_indices:
            continue
        part_no = _s(prow.get("PART_NO"))
        if not part_no:
            continue
        sec = _s(prow.get("SEC"))
        lvl1 = _s(prow.get("LVL1"))
        lvl2 = _s(prow.get("LVL2"))
        lvl3 = _s(prow.get("LVL3"))
        lvl4 = _s(prow.get("LVL4"))
        key = (sec, lvl1, lvl2, lvl3, lvl4, part_no)
        if key in seen:
            continue
        seen.add(key)

        part_name = _s(prow.get("PART_NAME"))
        qty = _s(prow.get("QUANTITY"))
        rows.append(TBRow(
            sec=sec, lvl1=lvl1, lvl2=lvl2, lvl3=lvl3, lvl4=lvl4,
            part_no=part_no, part_name=part_name, qty=qty,
        ))

    set_count = len(ps_groups)
    direct_count = len(rows) - sum(len(_expand_set_no(ps, ypl)) if "**" in ps else 1 for ps in ps_groups)
    logs.append(f"[STEP 4] 抽取完成: {len(rows)} 行 (SET 组 {set_count} 个, 直通行 {len(rows) - sum(1 for _ in [])} 个)")
    for w in set_warnings:
        logs.append(f"  [WARN] {w}")
    return rows, logs


def _expand_set_no(ps: str, ypl: pd.DataFrame) -> list[tuple[str, str]]:
    """展开含 -** 的 PART SET NO, 返回 [(完整零件号, 色型标签)]."""
    if ypl.empty or "MANUFACTURING_ASSY" not in ypl.columns or "CT_COMMENT" not in ypl.columns:
        return []
    mask = ypl["MANUFACTURING_ASSY"].astype(str).str.strip() == ps.strip()
    matches = ypl[mask]
    for comment in matches["CT_COMMENT"].dropna().unique():
        result = _parse_ct_comment(str(comment))
        if result:
            return result
    return []


def _format_assy_name(name: str) -> str:
    """Section A: 加 'ASSY.' 后缀."""
    if not name:
        return ""
    name = name.strip()
    if "ASSY" in name.upper():
        return name
    return f"{name} ASSY."


def step5_supplier(rows: list[TBRow], cu_fa_map: dict[str, str]) -> tuple[list[TBRow], list[str]]:
    """STEP 5: SUPPLIER 映射. CU_FA='A2' → 'YMC'. 显示: use_in != F3 → 'YMC TO F3'."""
    logs = []
    supplier_name = cu_fa_map.get("A2", "YMC")
    for row in rows:
        if row.use_in and row.use_in.upper() != "F3":
            row.supplier = f"{supplier_name} TO F3"
        else:
            row.supplier = supplier_name
    logs.append(f"[STEP 5] SUPPLIER 映射: CU_FA A2 → {supplier_name}")
    return rows, logs


def step6_ex_new(rows: list[TBRow], ln_bom_set: set[str]) -> tuple[list[TBRow], list[str]]:
    """STEP 6: EX/NEW 判定. PART_NO 在 LN BOM → EX, 否则 → NEW."""
    logs = []
    ex_count = 0
    new_count = 0
    for row in rows:
        pn = row.part_no.strip()
        if pn in ln_bom_set:
            row.ex_new = "EX"
            ex_count += 1
        else:
            row.ex_new = "NEW"
            new_count += 1
    logs.append(f"[STEP 6] EX/NEW 判定: EX={ex_count}, NEW={new_count}")
    return rows, logs


def step7_use_in(
    rows: list[TBRow],
    qcps_sub: pd.DataFrame,
    qcps_mlr: set[str],
    qcps_mll: set[str],
    email_map: dict[str, str],
) -> tuple[list[TBRow], list[str]]:
    """STEP 7: USE IN 判定 (QCPS 级联查询).

    SOP 规则:
    1. QCPS Sub Assy + STATION = EXTERNAL VENDOR → USE IN = EMAIL 供应商名
    2. QCPS Sub Assy + STATION = SUB ASSY → USE IN = F3
    3. QCPS ML(R) → F3
    4. QCPS ML(L) → F3
    5. 默认 → F3
    """
    logs = []

    # 构建 Sub Assy 索引: PART_NO → (STATION, PROCESS_NAME)
    sub_idx: dict[str, dict] = {}
    if not qcps_sub.empty and "PART_NO" in qcps_sub.columns:
        for _, r in qcps_sub.iterrows():
            pn = str(r.get("PART_NO", "")).strip()
            if not pn:
                continue
            station = str(r.get("STATION", "")).strip().upper()
            process = str(r.get("PROCESS_NAME", "")).strip()
            if pn not in sub_idx:
                sub_idx[pn] = {"STATION": station, "PROCESS_NAME": process}
            elif "EXTERNAL VENDOR" in station and "EXTERNAL VENDOR" not in sub_idx[pn]["STATION"]:
                sub_idx[pn] = {"STATION": station, "PROCESS_NAME": process}

    f3_count = 0
    vendor_count = 0
    for row in rows:
        pn = row.part_no.strip()
        use_in = "F3"

        if pn in sub_idx:
            info = sub_idx[pn]
            station = info["STATION"]
            if "EXTERNAL VENDOR" in station:
                use_in = email_map.get(pn, "MODERNRIA")
                vendor_count += 1
            else:
                use_in = "F3"
                f3_count += 1
        elif pn in qcps_mlr:
            use_in = "F3"
            f3_count += 1
        elif pn in qcps_mll:
            use_in = "F3"
            f3_count += 1
        else:
            f3_count += 1

        row.use_in = use_in

    logs.append(f"[STEP 7] USE IN 判定: F3={f3_count}, External Vendor={vendor_count}")

    # SUPPLIER 需要根据 USE IN 重新格式化
    for row in rows:
        supplier_name = "YMC"
        if row.use_in and row.use_in.upper() != "F3":
            row.supplier = f"{supplier_name} TO F3"
        else:
            row.supplier = supplier_name

    return rows, logs


# ---------------------------------------------------------------------------
# Part 3: 主生成函数
# ---------------------------------------------------------------------------

def generate_section_a(
    pcl: pd.DataFrame,
    cu_fa_map: dict[str, str],
    ln_bom_set: set[str],
    qcps_sub: pd.DataFrame,
    qcps_mlr: set[str],
    qcps_mll: set[str],
    ypl: pd.DataFrame,
    email_map: dict[str, str],
) -> tuple[list[TBRow], list[str]]:
    """主生成函数. 严格按 SOP STEP 1-7 执行."""
    all_logs: list[str] = []

    df, logs = step1_filter_mpl(pcl)
    all_logs.extend(logs)

    df, logs = step2_filter_engine(df)
    all_logs.extend(logs)

    df, logs = step3_filter_ymc(df)
    all_logs.extend(logs)

    if df.empty:
        all_logs.append("[ERROR] 过滤后无数据")
        return [], all_logs

    rows, logs = step4_extract_info(df, ypl)
    all_logs.extend(logs)

    rows, logs = step5_supplier(rows, cu_fa_map)
    all_logs.extend(logs)

    rows, logs = step6_ex_new(rows, ln_bom_set)
    all_logs.extend(logs)

    rows, logs = step7_use_in(rows, qcps_sub, qcps_mlr, qcps_mll, email_map)
    all_logs.extend(logs)

    return rows, all_logs


# ---------------------------------------------------------------------------
# Part 4: Excel Writer
# ---------------------------------------------------------------------------

COL_HEADERS = [
    "No.", "SEC", "LVL1", "LVL2", "LVL3", "LVL4",
    "PART NO.", "PART NAME", "QTY/ BIKE", "SUPPLIER", "USE IN", "EX/NEW", "REMARKS",
]
COL_WIDTHS = [6, 8, 8, 8, 8, 8, 22, 28, 10, 18, 14, 10, 18]

THIN = Side(style="thin", color="000000")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
FONT_TITLE = Font(name="Arial", size=14, bold=True, color="FFFFFF")
FONT_HEADER = Font(name="Arial", size=11, bold=True)
FONT_SECTION = Font(name="Arial", size=11, bold=True, color="FFFFFF")
FONT_COL = Font(name="Arial", size=10, bold=True)
FONT_DATA = Font(name="Arial", size=10)
FILL_TITLE = PatternFill("solid", fgColor="1F4E78")
FILL_SECTION = PatternFill("solid", fgColor="2E75B6")
FILL_COL = PatternFill("solid", fgColor="D9E1F2")
FILL_META = PatternFill("solid", fgColor="F2F2F2")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def write_section_a_excel(rows: list[TBRow], tb_no: str, target_factory: str, model: str, model_name: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = f"{tb_no}-SectionA"[:31]

    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    r = 1
    # TB 抬头
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=13)
    cell = ws.cell(r, 1, value="HLYM TECHNICAL BULLETIN")
    cell.font = FONT_TITLE
    cell.fill = FILL_TITLE
    cell.alignment = CENTER
    ws.row_dimensions[r].height = 28
    r += 1

    # 元信息
    meta = [
        ("TB NO.", f"{tb_no} (REV. 0)"),
        ("MODEL", f"{model} - {model_name}"),
        ("PURPOSE", f"{model} ({model_name}) - MASTER PART LIST FOR FACTORY {target_factory.lstrip('F')}"),
        ("WITH EFFECTIVE FROM", "TBD"),
    ]
    for label, value in meta:
        ws.cell(r, 1, value=label).font = FONT_HEADER
        ws.cell(r, 1).fill = FILL_META
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=13)
        ws.cell(r, 2, value=value).alignment = LEFT
        r += 1

    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=13)
    ws.cell(r, 1, value="The details are as listed below:").font = FONT_HEADER
    r += 2

    # Section A header
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=13)
    sec_cell = ws.cell(r, 1, value=f"A. PARTS WHICH SUPPLY FROM YMC TO {target_factory}")
    sec_cell.font = FONT_SECTION
    sec_cell.fill = FILL_SECTION
    sec_cell.alignment = LEFT
    ws.row_dimensions[r].height = 22
    r += 1

    # Column headers
    for c, h in enumerate(COL_HEADERS, 1):
        cell = ws.cell(r, c, value=h)
        cell.font = FONT_COL
        cell.fill = FILL_COL
        cell.alignment = CENTER
        cell.border = BORDER
    r += 1

    # Data rows
    for i, row in enumerate(rows, 1):
        vals = [i, row.sec, row.lvl1, row.lvl2, row.lvl3, row.lvl4,
                row.part_no, row.part_name, row.qty,
                row.supplier, row.use_in, row.ex_new, row.remarks]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(r, c, value=v)
            cell.font = FONT_DATA
            cell.alignment = LEFT if c in (8, 10, 11) else CENTER
            cell.border = BORDER
        r += 1

    ws.freeze_panes = "A9"

    import io
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ---------------------------------------------------------------------------
# Part 5: Verification
# ---------------------------------------------------------------------------

def verify_section_a(rows: list[TBRow], truth: list[dict]) -> dict:
    gen = [
        {"sec": r.sec, "lvl1": r.lvl1, "part_no": r.part_no, "part_name": r.part_name,
         "qty": r.qty, "supplier": r.supplier, "use_in": r.use_in, "ex_new": r.ex_new, "remarks": r.remarks}
        for r in rows
    ]
    def keyof(r):
        return (r.get("sec", ""), r.get("lvl1", ""), r.get("part_no", ""))

    truth_keys = {keyof(r): r for r in truth}
    gen_keys = {keyof(r): r for r in gen}

    matched = 0
    field_mismatches = []
    missing = []
    extra = []

    for key, t in truth_keys.items():
        if key in gen_keys:
            matched += 1
            g = gen_keys[key]
            diff = {}
            for f in ("part_name", "qty", "supplier", "use_in", "ex_new"):
                tv = (t.get(f) or "").strip()
                gv = (g.get(f) or "").strip()
                if tv.lower() != gv.lower():
                    diff[f] = {"truth": tv, "generated": gv}
            if diff:
                field_mismatches.append({"key": key, "part_no": t.get("part_no"), "diff": diff})
        else:
            missing.append(t)

    for key, g in gen_keys.items():
        if key not in truth_keys:
            extra.append(g)

    rate = matched / len(truth) * 100 if truth else 0
    return {
        "truth_rows": len(truth), "gen_rows": len(gen),
        "matched": matched, "rate": rate,
        "missing": missing, "extra": extra,
        "field_mismatches": field_mismatches,
    }


# ---------------------------------------------------------------------------
# Part 6: Streamlit UI
# ---------------------------------------------------------------------------

st.set_page_config(
    page_title="TB for YMC (Section A)",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.title("TB for YMC — Section A Auto Generator")
st.caption("严格按 SOP 文档 'TB for YMC' 实现: STEP 1-3 过滤 → STEP 4-7 填表 → SITUATION 2/3 SET NO 处理")

# --- Sidebar ---
st.sidebar.header("Data Source")
src_mode = st.sidebar.radio("Source", ["Built-in sample data", "Upload custom Excel"], index=0)

struct_arg: Optional[Path] = None

if src_mode == "Built-in sample data":
    struct_arg = STRUCT_PATH
    if struct_arg.exists():
        st.sidebar.success(f"Using: {struct_arg.name}")
    else:
        st.sidebar.error(f"File not found:\n{struct_arg}")
        st.stop()
else:
    upload_struct = st.sidebar.file_uploader("TB Structure Explanation.xlsx", type=["xlsx"], key="struct")
    if upload_struct:
        tmp_dir = Path(tempfile.mkdtemp())
        struct_arg = tmp_dir / upload_struct.name
        struct_arg.write_bytes(upload_struct.read())
        st.sidebar.success("File uploaded")
    else:
        st.sidebar.info("Please upload the Excel file")
        st.stop()

st.sidebar.divider()
st.sidebar.header("TB Parameters")
tb_no = st.sidebar.text_input("TB No.", value="TB-25-137")
target_factory = st.sidebar.selectbox("Target Factory", ["F1", "F2", "F3", "F4", "F5"], index=2)
run_btn = st.sidebar.button("Generate Section A", type="primary", use_container_width=True)

if not run_btn:
    st.info("Select data source on the left, then click **Generate Section A**.")
    st.stop()

# --- Load Data ---
with st.status("Loading source data...", expanded=True) as status:
    try:
        pcl = load_pcl(struct_arg)
        cu_fa_map = load_cu_fa_code(struct_arg)
        ln_bom_set = load_ln_bom(struct_arg)
        qcps_sub = load_qcps_sub(struct_arg)
        qcps_mlr = load_qcps_ml(struct_arg, "4. QCPS - ML (R)")
        qcps_mll = load_qcps_ml(struct_arg, "5. QCPS - ML (L)")
        ypl = load_ypl(struct_arg)
        email_map = load_email(struct_arg)
    except Exception as e:
        status.update(label=f"Load failed: {e}", state="error")
        st.exception(e)
        st.stop()

    status.write(f"PCL: **{len(pcl)}** rows")
    if "MPL_OUT_SIGN" in pcl.columns:
        vals = pcl["MPL_OUT_SIGN"].astype(str).str.strip().value_counts().to_dict()
        status.write(f"  MPL_OUT_SIGN: {vals}")
    else:
        status.write(f"  MPL_OUT_SIGN: (not present)")
    if "ENGINE_SIGN" in pcl.columns:
        vals = pcl["ENGINE_SIGN"].astype(str).str.strip().value_counts().to_dict()
        status.write(f"  ENGINE_SIGN: {vals}")
    else:
        status.write(f"  ENGINE_SIGN: (not present)")
    if "CU_FA" in pcl.columns:
        status.write(f"  CU_FA: {pcl['CU_FA'].value_counts().to_dict()}")
    status.write(f"CU_FA CODE: {len(cu_fa_map)} mappings (A2 = {cu_fa_map.get('A2', '?')})")
    status.write(f"LN BOM: {len(ln_bom_set)} parts")
    status.write(f"QCPS Sub Assy: {len(qcps_sub)} entries | ML(R): {len(qcps_mlr)} | ML(L): {len(qcps_mll)}")
    status.write(f"YPL: {len(ypl)} rows | EMAIL: {len(email_map)} supplier mappings")
    model = str(pcl.get("MODEL", pd.Series([""])).iloc[0]) if not pcl.empty else ""
    model_name = str(pcl.get("MODEL_NAME", pd.Series([""])).iloc[0]) if not pcl.empty and "MODEL_NAME" in pcl.columns else ""
    status.write(f"Model: **{model}** ({model_name})")
    status.update(label="Data loaded", state="complete")

# --- Generate ---
with st.status("Applying SOP rules...", expanded=True) as status:
    rows, logs = generate_section_a(pcl, cu_fa_map, ln_bom_set, qcps_sub, qcps_mlr, qcps_mll, ypl, email_map)
    for log in logs:
        status.write(log)
    status.update(label=f"Section A generated: {len(rows)} rows", state="complete")

# --- Metrics ---
ex_count = sum(1 for r in rows if r.ex_new == "EX")
new_count = sum(1 for r in rows if r.ex_new == "NEW")
f3_count = sum(1 for r in rows if r.use_in == "F3")
vendor_count = sum(1 for r in rows if r.use_in != "F3")

cols = st.columns(5)
cols[0].metric("Total Parts", len(rows))
cols[1].metric("EX", ex_count)
cols[2].metric("NEW", new_count)
cols[3].metric("USE IN = F3", f3_count)
cols[4].metric("USE IN = Vendor", vendor_count)

# --- SOP Steps Detail ---
with st.expander("SOP Step Details", expanded=False):
    for log in logs:
        if "[ERROR]" in log:
            st.error(log)
        elif "[WARN]" in log:
            st.warning(log)
        elif "[SKIP]" in log:
            st.info(log)
        else:
            st.write(log)

# --- Result Table ---
st.subheader(f"Section A: PARTS WHICH SUPPLY FROM YMC TO {target_factory}")
if rows:
    rows_data = [
        {"No.": i, "SEC": r.sec, "LVL1": r.lvl1, "LVL2": r.lvl2, "LVL3": r.lvl3, "LVL4": r.lvl4,
         "PART NO.": r.part_no, "PART NAME": r.part_name, "QTY/BIKE": r.qty,
         "SUPPLIER": r.supplier, "USE IN": r.use_in, "EX/NEW": r.ex_new, "REMARKS": r.remarks}
        for i, r in enumerate(rows, 1)
    ]
    st.dataframe(pd.DataFrame(rows_data), use_container_width=True, hide_index=True)
else:
    st.warning("No Section A rows generated.")

# --- Download ---
st.divider()
st.subheader("Download")

xlsx_bytes = write_section_a_excel(rows, tb_no, target_factory, model, model_name)
st.download_button(
    label="Download Section_A.xlsx",
    data=xlsx_bytes,
    file_name=f"{tb_no}-{target_factory}-SectionA.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    type="primary",
    use_container_width=True,
)